"""Quick standalone analysis report — Excel + Markdown.

Runs the screening pipeline on the watchlist and produces a comprehensive
multi-sheet Excel workbook for manual analysis. No portfolio, no Telegram.

Usage:
    python quick_analysis.py                    # watchlist (default)
    python quick_analysis.py --india            # India only
    python quick_analysis.py --us               # US only
    python quick_analysis.py --symbols RELIANCE.NS,NVDA,TCS.NS

Output:
    reports/analysis_<YYYY-MM-DD>.xlsx
    reports/analysis_<YYYY-MM-DD>.md
"""
from __future__ import annotations
import argparse
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from tqdm import tqdm

from config import REPORTS_DIR, WATCHLIST_INDIA, WATCHLIST_US
from data_sources.yahoo import fetch_many
from analysis.composite import analyze, StockReport
from analysis.indicators import atr, annualized_volatility

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("quick")


# ── Row builder ──────────────────────────────────────────────────────────────
def _row(r: StockReport, td) -> dict:
    """Flatten a StockReport into a single Excel row with derived trade levels."""
    price = r.price or 0.0
    a = atr(td.history) if td and td.ok else None
    vol = annualized_volatility(td.history) if td and td.ok else None
    atr_pct = (a / price * 100) if (a and price) else None

    # Suggested levels (matches portfolio defaults)
    stop = None
    t1 = None
    t2 = None
    if price > 0:
        atr_stop = price - 2.5 * a if a else None
        hard_stop = price * 0.85
        stop = max(atr_stop, hard_stop) if atr_stop else hard_stop
        t1 = price * 1.22
        t2 = price * 1.38

    f = r.fundamental
    m = r.momentum
    t = r.technical

    return {
        "Symbol": r.symbol,
        "Name": (r.name or "")[:40],
        "Market": r.market,
        "Sector": r.sector or "",
        "Price": round(price, 2),
        "Score": round(r.composite_score, 1),
        "Verdict": r.verdict,
        # Returns
        "Ret_1W_%": _pct(m.get("ret_1w")),
        "Ret_1M_%": _pct(m.get("ret_1m")),
        "Ret_3M_%": _pct(m.get("ret_3m")),
        "Ret_6M_%": _pct(m.get("ret_6m")),
        "Ret_1Y_%": _pct(m.get("ret_1y")),
        # Technicals
        "RSI": _round(t.get("rsi"), 1),
        "Above_50DMA": t.get("above_sma50"),
        "Above_200DMA": t.get("above_sma200"),
        "From_52W_High_%": _pct(t.get("pct_from_52w_high")),
        "ATR_%": _round(atr_pct, 2),
        "Ann_Vol_%": _round((vol or 0) * 100, 1) if vol else None,
        # Fundamentals
        "P/E": _round(f.get("pe"), 1),
        "P/B": _round(f.get("pb"), 2),
        "ROE_%": _round((f.get("roe") or 0) * 100, 1) if f.get("roe") is not None else None,
        "D/E": _round(f.get("debt_to_equity"), 1),
        "EPS_Growth_%": _round((f.get("eps_growth") or 0) * 100, 1) if f.get("eps_growth") is not None else None,
        "Mkt_Cap_Cr": _round((f.get("market_cap") or 0) / 1e7, 0) if r.market == "IN" else _round((f.get("market_cap") or 0) / 1e9, 1),
        # Component scores
        "Tech_Score": _round(t.get("score"), 0),
        "Fund_Score": _round(f.get("score"), 0),
        "Momentum_Score": _round(m.get("score"), 0),
        "Sentiment_Score": _round(r.sentiment.get("score"), 0),
        "Forecast_Score": _round(r.forecast.get("score"), 0),
        # Suggested trade levels
        "Suggested_Entry": _round(price, 2),
        "Stop_Loss": _round(stop, 2),
        "Stop_Loss_%": _round(((stop / price) - 1) * 100, 1) if (stop and price) else None,
        "Target_T1_+20%": _round(t1, 2),
        "Target_T2_+35%": _round(t2, 2),
        # Top signals
        "Key_Signals": " | ".join(r.all_signals[:4]),
    }


def _pct(v):
    return _round((v or 0) * 100, 1) if v is not None else None


def _round(v, n=2):
    try:
        return round(float(v), n) if v is not None else None
    except (TypeError, ValueError):
        return None


# ── Excel writer with formatting ─────────────────────────────────────────────
def _write_excel(df: pd.DataFrame, path: Path) -> None:
    """Write multi-sheet Excel with conditional formatting."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: All ranked
        df_sorted = df.sort_values("Score", ascending=False)
        df_sorted.to_excel(writer, sheet_name="All_Ranked", index=False)

        # Sheet 2: Top picks (Score >= 62, BUY+)
        top = df_sorted[df_sorted["Score"] >= 62]
        if not top.empty:
            top.to_excel(writer, sheet_name="Top_Picks", index=False)

        # Sheet 3: India only
        india = df_sorted[df_sorted["Market"] == "IN"]
        if not india.empty:
            india.to_excel(writer, sheet_name="India", index=False)

        # Sheet 4: US only
        us = df_sorted[df_sorted["Market"] == "US"]
        if not us.empty:
            us.to_excel(writer, sheet_name="US", index=False)

        # Sheet 5: By sector (avg score per sector)
        if "Sector" in df.columns and df["Sector"].notna().any():
            sector = (df[df["Sector"] != ""]
                      .groupby(["Market", "Sector"])
                      .agg(Count=("Symbol", "count"),
                           Avg_Score=("Score", "mean"),
                           Avg_1M=("Ret_1M_%", "mean"),
                           Avg_3M=("Ret_3M_%", "mean"),
                           Avg_RSI=("RSI", "mean"))
                      .round(1)
                      .reset_index()
                      .sort_values("Avg_Score", ascending=False))
            sector.to_excel(writer, sheet_name="By_Sector", index=False)

        # Sheet 6: AVOID list (Score < 32)
        avoid = df_sorted[df_sorted["Score"] < 45]
        if not avoid.empty:
            avoid.to_excel(writer, sheet_name="Avoid", index=False)

    # Post-format with openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import ColorScaleRule

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "B2"

        # Auto width (capped)
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0)
                          for c in col_cells)
            letter = col_cells[0].column_letter
            ws.column_dimensions[letter].width = min(max_len + 2, 32)

        # Color-scale on Score column (red→yellow→green)
        score_col = None
        for cell in ws[1]:
            if cell.value == "Score":
                score_col = cell.column_letter
                break
        if score_col and ws.max_row > 1:
            rng = f"{score_col}2:{score_col}{ws.max_row}"
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(start_type="num", start_value=20, start_color="F8696B",
                               mid_type="num", mid_value=55, mid_color="FFEB84",
                               end_type="num", end_value=85, end_color="63BE7B"),
            )

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# ── Markdown summary (concise) ───────────────────────────────────────────────
def _write_markdown(df: pd.DataFrame, path: Path) -> None:
    today = datetime.now().strftime("%Y-%m-%d")
    lines = [f"# Stock Analysis — {today}\n"]
    lines.append(f"**Universe:** {len(df)} stocks "
                 f"({(df['Market']=='IN').sum()} IN, {(df['Market']=='US').sum()} US)\n")

    avg = df["Score"].mean()
    lines.append(f"**Average score:** {avg:.1f}\n")

    counts = df["Verdict"].value_counts().to_dict()
    lines.append("**Distribution:** " + ", ".join(f"{k}: {v}" for k, v in counts.items()) + "\n")

    # Top 15
    top = df.sort_values("Score", ascending=False).head(15)
    lines.append("\n## Top 15 Picks\n")
    cols = ["Symbol", "Market", "Score", "Verdict", "Price",
            "Ret_1M_%", "Ret_3M_%", "RSI", "P/E", "Sector"]
    lines.append(top[cols].to_markdown(index=False))

    # Sector heat
    if (df["Sector"] != "").any():
        sec = (df[df["Sector"] != ""]
               .groupby(["Market", "Sector"])
               .agg(N=("Symbol", "count"), AvgScore=("Score", "mean"))
               .round(1).reset_index()
               .sort_values("AvgScore", ascending=False).head(15))
        lines.append("\n## Top Sectors by Avg Score\n")
        lines.append(sec.to_markdown(index=False))

    # Avoid list
    avoid = df[df["Score"] < 45].sort_values("Score").head(10)
    if not avoid.empty:
        lines.append("\n## ⚠️ Avoid (Score < 45)\n")
        lines.append(avoid[["Symbol", "Market", "Score", "Verdict",
                            "Ret_1M_%", "RSI", "Sector"]].to_markdown(index=False))

    lines.append("\n---\n*Not investment advice. Cross-check fundamentals on screener.in / finviz.com before acting.*\n")
    path.write_text("\n".join(lines), encoding="utf-8")


# ── Main ─────────────────────────────────────────────────────────────────────
def run(symbols: list[str]) -> None:
    log.info("Universe: %d symbols", len(symbols))
    data = fetch_many(symbols, period="1y")

    rows = []
    for sym, td in tqdm(data.items(), desc="Analyzing"):
        try:
            r = analyze(td)
            if r.composite_score > 0:
                rows.append(_row(r, td))
        except Exception as e:
            log.warning("Skip %s: %s", sym, e)

    if not rows:
        log.error("No data — likely network issue (yfinance blocked?)")
        return

    df = pd.DataFrame(rows)
    today = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = REPORTS_DIR / f"analysis_{today}.xlsx"
    md_path = REPORTS_DIR / f"analysis_{today}.md"

    _write_excel(df, xlsx_path)
    _write_markdown(df, md_path)

    log.info("✅ Excel:    %s", xlsx_path)
    log.info("✅ Markdown: %s", md_path)
    print(f"\n📊 Analyzed {len(df)} stocks")
    print(f"📁 Excel:    {xlsx_path}")
    print(f"📄 Markdown: {md_path}\n")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Quick stock analysis report")
    p.add_argument("--india", action="store_true", help="India watchlist only")
    p.add_argument("--us", action="store_true", help="US watchlist only")
    p.add_argument("--symbols", help="Comma-separated tickers (override watchlist)")
    args = p.parse_args()

    if args.symbols:
        syms = [s.strip() for s in args.symbols.split(",") if s.strip()]
    elif args.india:
        syms = WATCHLIST_INDIA
    elif args.us:
        syms = WATCHLIST_US
    else:
        syms = WATCHLIST_INDIA + WATCHLIST_US

    run(syms)
