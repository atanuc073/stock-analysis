"""Generate Markdown + HTML + Excel reports from ranked StockReport list."""
from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, TOP_PICKS_SECTOR_CAP


def _fmt_money(v):
    if v is None or v == 0:
        return "—"
    if v >= 1e12: return f"${v/1e12:.2f}T"
    if v >= 1e9: return f"${v/1e9:.2f}B"
    if v >= 1e6: return f"${v/1e6:.0f}M"
    return f"${v:,.0f}"


def _fmt_pct(v, signed=True):
    if v is None:
        return "—"
    return f"{'+' if signed and v >= 0 else ''}{v:.2f}%"


def _score_of(r) -> float:
    """Prefer adjusted_score (cross-sectional pass) if present, else composite."""
    return getattr(r, "adjusted_score", None) or getattr(r, "composite_score", 0.0)


def _select_top_picks(reports: list, top_n: int, sector_cap: int = TOP_PICKS_SECTOR_CAP) -> list:
    """Pick top-N by score with a per-sector cap to avoid sector concentration.

    Walks the score-sorted list and adds each name unless its sector has
    already reached `sector_cap` slots. Names skipped over are still kept as
    overflow — if we can't fill `top_n` with the cap enforced (e.g. sparse
    universe), the overflow is appended in score order to guarantee a full
    list. `sector_cap <= 0` or `>= top_n` disables the cap.
    """
    sorted_rs = sorted(reports, key=_score_of, reverse=True)
    if sector_cap <= 0 or sector_cap >= top_n:
        return sorted_rs[:top_n]
    picks: list = []
    overflow: list = []
    counts: dict[str, int] = {}
    for r in sorted_rs:
        sec = (getattr(r, "sector", "") or "Unknown")
        if counts.get(sec, 0) < sector_cap:
            picks.append(r)
            counts[sec] = counts.get(sec, 0) + 1
            if len(picks) >= top_n:
                break
        else:
            overflow.append(r)
    if len(picks) < top_n:
        picks.extend(overflow[: top_n - len(picks)])
    return picks


def render_markdown(reports: list, top_n: int = 15) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    sorted_rs = sorted(reports, key=_score_of, reverse=True)
    top = _select_top_picks(reports, top_n)
    bottom = sorted_rs[-min(5, len(sorted_rs)):]

    lines = []
    lines.append(f"# Daily Stock Analysis — {today}\n")
    lines.append(f"_Universe: {len(reports)} tickers analyzed_\n")

    # Summary table
    lines.append("## Top Picks\n")
    lines.append("| # | Ticker | Mkt | Name | Price | Stop | T1 (+22%) | Score | Verdict | 1D | 1M | 3M | RSI | P/E | Sector |")
    lines.append("|---|--------|-----|------|-------|------|-----------|-------|---------|------|------|------|-----|-----|--------|")
    for i, r in enumerate(top, 1):
        pe_val = r.fundamental.get('pe')
        pe_str = f"{pe_val:.1f}" if isinstance(pe_val, float) else (pe_val or '—')
        atr_val = r.technical.get('atr', 0.0)
        atr_stop = r.price - 3.0 * atr_val if atr_val > 0 else r.price * 0.82
        stop_price = max(atr_stop, r.price * 0.82)
        t1_price = r.price * 1.22
        lines.append(
            f"| {i} | **{r.symbol}** | {r.market} | {(r.name or '')[:24]} | "
            f"{r.price:.2f} | {stop_price:.2f} | {t1_price:.2f} | **{r.composite_score:.1f}** | {r.verdict} | "
            f"{_fmt_pct(r.momentum.get('ret_1d'))} | "
            f"{_fmt_pct(r.momentum.get('ret_1m'))} | {_fmt_pct(r.momentum.get('ret_3m'))} | "
            f"{r.technical.get('rsi', 0):.0f} | "
            f"{pe_str} | "
            f"{(r.sector or '')[:14]} |"
        )

    lines.append("\n## Detailed Analysis (Top Picks)\n")
    for r in top:
        atr_val = r.technical.get('atr', 0.0)
        atr_stop = r.price - 3.0 * atr_val if atr_val > 0 else r.price * 0.82
        stop_price = max(atr_stop, r.price * 0.82)
        stop_pct = (stop_price / r.price - 1) * 100 if r.price > 0 else -18.0
        t1_price = r.price * 1.22
        lines.append(f"### {r.symbol} — {r.name} ({r.market})  →  **{r.verdict}** (score {r.composite_score:.1f})\n")
        lines.append(f"- **Price:** {r.price:.2f}  |  **Stop:** {stop_price:.2f} ({stop_pct:.1f}%)  |  **T1:** {t1_price:.2f} (+22%)  |  **Sector:** {r.sector or '—'}  |  **Mkt Cap:** {_fmt_money(r.fundamental.get('market_cap'))}")
        lines.append(f"- **Returns:** 1D {_fmt_pct(r.momentum.get('ret_1d'))}, 1W {_fmt_pct(r.momentum.get('ret_1w'))}, 1M {_fmt_pct(r.momentum.get('ret_1m'))}, 3M {_fmt_pct(r.momentum.get('ret_3m'))}, 1Y {_fmt_pct(r.momentum.get('ret_1y'))}")
        lines.append(f"- **Technical:** RSI {r.technical.get('rsi', 0):.1f} | Above 50DMA: {r.technical.get('above_sma50')} | Above 200DMA: {r.technical.get('above_sma200')} | {_fmt_pct(r.technical.get('pct_from_52w_high'), False)} from 52W high")
        f = r.fundamental
        lines.append(
            f"- **Fundamentals:** P/E {f.get('pe') or '—'} | P/B {f.get('pb') or '—'} | "
            f"ROE {(f.get('roe') or 0)*100:.1f}% | D/E {f.get('debt_to_equity') or '—'} | "
            f"EPS growth {(f.get('eps_growth') or 0)*100:.1f}%"
        )
        fc = r.forecast
        if fc.get("expected_return_pct") is not None:
            lines.append(f"- **Forecast (21d):** {_fmt_pct(fc.get('expected_return_pct'))} (target ~{fc.get('forecast_price'):.2f})")
        if r.options.get("available"):
            lines.append(f"- **Options (US):** PCR {r.options.get('put_call_ratio')} | calls {r.options.get('call_volume'):,} | puts {r.options.get('put_volume'):,} | exp {r.options.get('expiry')}")
        if r.sentiment.get("headlines"):
            lines.append(f"- **Sentiment:** avg {r.sentiment.get('avg_sentiment', 0):+.2f}")
            for h in r.sentiment["headlines"][:3]:
                lines.append(f"  - _{h['title']}_  ({h['sentiment']:+.2f})")
        if r.all_signals:
            lines.append(f"- **Signals:** {' • '.join(r.all_signals[:8])}")
        lines.append("")

    lines.append("## Avoid / Weakest\n")
    lines.append("| Ticker | Score | Verdict | Reason |")
    lines.append("|--------|-------|---------|--------|")
    for r in bottom:
        reason = (r.all_signals[0] if r.all_signals else r.error or "weak signals")
        lines.append(f"| {r.symbol} | {r.composite_score:.1f} | {r.verdict} | {reason} |")

    lines.append("\n---\n_⚠️ Disclaimer: This is automated technical/quantitative analysis, not investment advice. Always do your own due diligence and consider consulting a SEBI-registered advisor (India) or licensed financial advisor (US)._\n")
    return "\n".join(lines)


# ── Excel report ─────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT = Font(name="Calibri", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_NUM_FMT_PCT = '0.00"%"'
_NUM_FMT_SCORE = '0.0'
_NUM_FMT_PRICE = '#,##0.00'


def _score_fill(score: float) -> PatternFill:
    """Green (high) → Yellow (mid) → Red (low) gradient for 0-100 scores."""
    if score >= 70:
        return PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # green
    if score >= 62:
        return PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")  # light green
    if score >= 55:
        return PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")  # yellow
    if score >= 45:
        return PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # orange
    if score >= 35:
        return PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # red
    return PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")      # dark red


def _verdict_fill(verdict: str) -> PatternFill:
    colors = {
        "STRONG BUY": "1ABC9C",
        "BUY": "2ECC71",
        "HOLD": "F1C40F",
        "REDUCE": "E67E22",
        "AVOID": "E74C3C",
    }
    c = colors.get(verdict, "FFFFFF")
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def _risk_flags(r) -> str:
    """Compact risk indicators for quick filtering."""
    flags = []
    t = r.technical or {}
    m = r.momentum or {}
    ed = r.earnings_drift or {}
    rsi = t.get("rsi") or 0
    pct_from_high = t.get("pct_from_52w_high")
    vol_ratio = t.get("volume_ratio") or 1.0
    if t.get("extended_at_high") or rsi >= 75 or (pct_from_high is not None and pct_from_high > -2):
        flags.append("EXT")
    if t.get("above_sma200") is False:
        flags.append("BEAR")
    r1w = m.get("ret_1w")
    if r1w is not None and r1w < -8:
        flags.append("KNIFE")
    if vol_ratio < 0.4:
        flags.append("THIN")
    days = ed.get("earnings_gap_days_ago")
    if days is not None and 0 <= days <= 7:
        flags.append("ERPOST")
    if rsi and rsi <= 30:
        flags.append("OS")  # oversold = potential opportunity
    return " ".join(flags)


def _swing_setup(r) -> dict | None:
    """Evaluate a ticker as a 90-day swing trade setup.

    Returns a dict with entry/stop/target/RR/conviction, or None if it does
    not meet baseline filters.
    """
    t = r.technical or {}
    m = r.momentum or {}
    f = r.fundamental or {}
    fc = r.forecast or {}
    q = r.quality or {}

    price = r.price or t.get("price") or 0
    if price <= 0:
        return None

    rsi = t.get("rsi") or 0
    above_50 = bool(t.get("above_sma50"))
    above_200 = bool(t.get("above_sma200"))
    pct_from_high = t.get("pct_from_52w_high")
    vol_ratio = t.get("volume_ratio") or 1.0
    score = r.composite_score or 0
    adj = r.adjusted_score or score
    flags = _risk_flags(r).split()

    # ── Hard filters (must pass all) ─────────────────────────────────
    if score < 60: return None
    if adj < 58: return None
    if not (above_50 and above_200): return None
    if not (38 <= rsi <= 68): return None
    if pct_from_high is None or not (-22 <= pct_from_high <= -3): return None
    if vol_ratio < 0.8: return None
    if any(bad in flags for bad in ("EXT", "BEAR", "KNIFE", "ERPOST", "THIN")):
        return None
    # avoid binary earnings risk if earnings hit very recently or imminent
    ed = r.earnings_drift or {}
    er_days = ed.get("earnings_gap_days_ago")
    if er_days is not None and er_days < 5:
        return None

    # ── Targets / stops (90-day horizon) ─────────────────────────────
    # Stop: ATR-based (3× ATR) with -18% hard floor — matches lifecycle.py
    atr_val = t.get("atr", 0.0)
    atr_stop = price - 3.0 * atr_val if atr_val > 0 else price * 0.82
    hard_stop = price * 0.82
    stop = round(max(atr_stop, hard_stop), 2)
    stop_pct = 1.0 - stop / price if price > 0 else 0.18

    # Target: blend of forecast and rule-of-thumb (12-18% over 90d)
    fc_target = fc.get("forecast_price") or 0
    fc_ret = fc.get("expected_return_pct") or 0
    # Annualize-ish: 21d forecast → 90d ≈ x4.3, capped at +25%
    target_ret_90d = max(0.10, min(fc_ret * 4 / 100 if fc_ret else 0.12, 0.25))
    target = round(price * (1 + target_ret_90d), 2)
    if fc_target and fc_target > price * 1.10:
        target = round(max(target, fc_target * 1.05), 2)  # extend if forecast is bullish

    rr = round((target - price) / max(price - stop, 0.01), 2)

    # ── Setup quality 0-100 (confluence-based) ───────────────────────
    sq = 50
    sq += min(20, (score - 60) * 1.0)            # composite tilt
    sq += min(8, (adj - 58) * 0.8)               # cross-sectional tilt
    if 45 <= rsi <= 60: sq += 6                  # ideal RSI band
    if -15 <= pct_from_high <= -5: sq += 6       # ideal pullback zone
    if vol_ratio >= 1.2: sq += 5
    if vol_ratio >= 1.5: sq += 3
    eps_g = f.get("eps_growth") or 0
    if eps_g > 0.10: sq += 4
    if eps_g > 0.20: sq += 3
    if (q.get("score") or 50) >= 60: sq += 4
    if (m.get("score") or 50) >= 65: sq += 4
    if rr >= 2.0: sq += 4
    if rr >= 3.0: sq += 3
    sq = max(0, min(100, sq))

    # Conviction tier
    if sq >= 78 and rr >= 2.0:
        conviction = "A"
        size_pct = 7.0
    elif sq >= 68 and rr >= 1.6:
        conviction = "B"
        size_pct = 5.0
    else:
        conviction = "C"
        size_pct = 3.0

    return {
        "entry": round(price, 2),
        "stop": stop,
        "target": target,
        "stop_pct": round(-stop_pct * 100, 1),
        "target_pct": round(target_ret_90d * 100, 1),
        "rr": rr,
        "setup_quality": round(sq, 0),
        "conviction": conviction,
        "size_pct": size_pct,
        "rsi": rsi,
        "vol_ratio": vol_ratio,
        "pct_from_high": pct_from_high,
        "er_days": er_days,
    }


def _auto_width(ws, min_width=8, max_width=30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            lengths.append(len(val))
        best = min(max(max(lengths, default=min_width) + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = best


def _style_header_row(ws, row_num=1):
    for cell in ws[row_num]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER


def render_excel(reports: list, top_n: int = 15) -> Workbook:
    """Build a multi-sheet Excel workbook from StockReport list."""
    wb = Workbook()
    sorted_rs = sorted(reports, key=_score_of, reverse=True)
    top = _select_top_picks(reports, top_n)
    bottom = sorted_rs[-min(5, len(sorted_rs)):]

    # ── Sheet 1: Dashboard ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    headers = [
        "#", "Ticker", "Market", "Name", "Sector", "Price",
        "Stop (ATR)", "T1 (+22%)",
        "Score", "Adj Score", "Verdict",
        "Tech", "Fund", "Mom", "Qual",
        "RSI", "% from 52WH", "Vol Ratio", ">200DMA",
        "1D %", "1W %", "1M %", "3M %",
        "P/E", "P/B", "ROE %", "D/E", "FCF Yield %",
        "EPS Gr %", "Rev Gr %", "Div Yield %",
        "ER Gap %", "Days Since ER",
        "Forecast 21d %",
        "Risk Flags", "Signals",
    ]
    ws.append(headers)
    _style_header_row(ws)

    score_col = headers.index("Score") + 1
    adj_col = headers.index("Adj Score") + 1
    verdict_col = headers.index("Verdict") + 1
    flags_col = headers.index("Risk Flags") + 1
    signals_col = headers.index("Signals") + 1

    for i, r in enumerate(sorted_rs, 1):
        f = r.fundamental or {}
        m = r.momentum or {}
        fc = r.forecast or {}
        t = r.technical or {}
        q = r.quality or {}
        ed = r.earnings_drift or {}
        fcf_y = q.get("fcf_yield")
        
        atr_val = t.get('atr', 0.0)
        atr_stop = r.price - 3.0 * atr_val if atr_val > 0 else r.price * 0.82
        stop_price = max(atr_stop, r.price * 0.82)
        
        row = [
            i,
            r.symbol,
            r.market,
            (r.name or "")[:28],
            (r.sector or "")[:20],
            round(r.price, 2),
            round(stop_price, 2),
            round(r.price * 1.22, 2),
            round(r.composite_score, 1),
            round(r.adjusted_score or r.composite_score, 1),
            r.verdict,
            round(t.get("score") or 0, 0),
            round(f.get("score") or 0, 0),
            round(m.get("score") or 0, 0),
            round(q.get("score") or 0, 0),
            round(t.get("rsi", 0), 1),
            round(t.get("pct_from_52w_high") or 0, 1),
            round(t.get("volume_ratio") or 0, 2),
            "Yes" if t.get("above_sma200") else "No",
            round(m.get("ret_1d") or 0, 2),
            round(m.get("ret_1w") or 0, 2),
            round(m.get("ret_1m") or 0, 2),
            round(m.get("ret_3m") or 0, 2),
            round(f.get("pe") or 0, 1) if f.get("pe") else None,
            round(f.get("pb") or 0, 2) if f.get("pb") else None,
            round((f.get("roe") or 0) * 100, 1),
            round(f.get("debt_to_equity") or 0, 1) if f.get("debt_to_equity") else None,
            round(fcf_y * 100, 2) if fcf_y is not None else None,
            round((f.get("eps_growth") or 0) * 100, 1),
            round((f.get("revenue_growth") or 0) * 100, 1),
            round(f.get("dividend_yield") or 0, 2),
            round((ed.get("earnings_gap_pct") or 0), 2) if ed.get("earnings_gap_pct") is not None else None,
            ed.get("earnings_gap_days_ago"),
            round(fc.get("expected_return_pct") or 0, 2),
            _risk_flags(r),
            "; ".join(r.all_signals[:5]),
        ]
        ws.append(row)
        row_num = i + 1  # header is row 1
        # Color-code Score cell
        score_cell = ws.cell(row=row_num, column=score_col)
        score_cell.fill = _score_fill(r.composite_score)
        score_cell.font = Font(name="Calibri", bold=True, size=10,
                               color="FFFFFF" if r.composite_score >= 55 or r.composite_score < 35 else "000000")
        # Color-code Adjusted Score cell
        adj_val = r.adjusted_score or r.composite_score
        adj_cell = ws.cell(row=row_num, column=adj_col)
        adj_cell.fill = _score_fill(adj_val)
        adj_cell.font = Font(name="Calibri", bold=True, size=10,
                             color="FFFFFF" if adj_val >= 55 or adj_val < 35 else "000000")
        # Color-code Verdict cell
        verdict_cell = ws.cell(row=row_num, column=verdict_col)
        verdict_cell.fill = _verdict_fill(r.verdict)
        verdict_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        # Color-code component sub-scores (Tech, Fund, Mom, Qual)
        for sub_col in range(headers.index("Tech") + 1, headers.index("Qual") + 2):
            sub_cell = ws.cell(row=row_num, column=sub_col)
            try:
                v = float(sub_cell.value) if sub_cell.value is not None else 0
                sub_cell.fill = _score_fill(v)
                sub_cell.font = Font(name="Calibri", size=9,
                                     color="FFFFFF" if v >= 55 or v < 35 else "000000")
            except (TypeError, ValueError):
                pass
        # Risk flags styling: red bold if any flags present
        flags_cell = ws.cell(row=row_num, column=flags_col)
        if flags_cell.value:
            flags_cell.font = Font(name="Calibri", bold=True, size=9, color="C0392B")
        # Style all cells in this row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = _BORDER
            if cell.font == Font():  # only set if not already styled
                cell.font = _DATA_FONT
            if col_idx == signals_col:
                cell.alignment = _LEFT
            else:
                cell.alignment = _CENTER

    # Freeze pane after Verdict column for easy scrolling
    ws.freeze_panes = ws.cell(row=2, column=verdict_col + 1).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _auto_width(ws)

    # ── Sheet 2: Detailed Analysis (Top Picks) ──────────────────────────
    ws2 = wb.create_sheet(title="Top Picks Detail")
    detail_headers = [
        "Ticker", "Name", "Market", "Verdict", "Score",
        "Price", "Stop (ATR)", "T1 (+22%)", "Sector", "Mkt Cap",
        "RSI", "Above 50DMA", "Above 200DMA", "% from 52W High",
        "P/E", "P/B", "ROE %", "D/E", "EPS Gr %", "Rev Gr %", "Profit Margin %",
        "1D %", "1W %", "1M %", "3M %",
        "Forecast 21d %", "Forecast Price",
        "Sentiment Avg", "Top Headline",
        "Signals",
    ]
    ws2.append(detail_headers)
    _style_header_row(ws2)

    for r in top:
        f = r.fundamental
        m = r.momentum
        fc = r.forecast
        t = r.technical
        s = r.sentiment
        headline = ""
        if s.get("headlines"):
            headline = s["headlines"][0].get("title", "")[:60]
            
        atr_val = t.get('atr', 0.0)
        atr_stop = r.price - 3.0 * atr_val if atr_val > 0 else r.price * 0.82
        stop_price = max(atr_stop, r.price * 0.82)
            
        row = [
            r.symbol,
            (r.name or "")[:28],
            r.market,
            r.verdict,
            round(r.composite_score, 1),
            round(r.price, 2),
            round(stop_price, 2),
            round(r.price * 1.22, 2),
            r.sector or "",
            _fmt_money(f.get("market_cap")),
            round(t.get("rsi", 0), 1),
            "Yes" if t.get("above_sma50") else "No",
            "Yes" if t.get("above_sma200") else "No",
            round(t.get("pct_from_52w_high", 0), 2),
            round(f.get("pe") or 0, 1) if f.get("pe") else None,
            round(f.get("pb") or 0, 2) if f.get("pb") else None,
            round((f.get("roe") or 0) * 100, 1),
            round(f.get("debt_to_equity") or 0, 1) if f.get("debt_to_equity") else None,
            round((f.get("eps_growth") or 0) * 100, 1),
            round((f.get("revenue_growth") or 0) * 100, 1),
            round((f.get("profit_margin") or 0) * 100, 1),
            round(m.get("ret_1d") or 0, 2),
            round(m.get("ret_1w") or 0, 2),
            round(m.get("ret_1m") or 0, 2),
            round(m.get("ret_3m") or 0, 2),
            round(fc.get("expected_return_pct") or 0, 2),
            round(fc.get("forecast_price") or 0, 2),
            round(s.get("avg_sentiment", 0), 3),
            headline,
            "; ".join(r.all_signals[:6]),
        ]
        ws2.append(row)

    # Style detail sheet
    for row_cells in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(detail_headers)):
        for cell in row_cells:
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER
        # Color score
        score_cell = row_cells[4]  # column E = Score
        if score_cell.value is not None:
            score_cell.fill = _score_fill(score_cell.value)
            score_cell.font = Font(name="Calibri", bold=True, size=10,
                                   color="FFFFFF" if score_cell.value >= 55 or score_cell.value < 35 else "000000")
        verdict_cell = row_cells[3]  # column D = Verdict
        if verdict_cell.value:
            verdict_cell.fill = _verdict_fill(verdict_cell.value)
            verdict_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}1"
    _auto_width(ws2)

    # ── Sheet 3: Swing 90D Scanner ──────────────────────────────────────
    ws_sw = wb.create_sheet(title="Swing 90D")
    swing_headers = [
        "Rank", "Ticker", "Market", "Name", "Sector",
        "Conviction", "Setup Q", "Score", "Adj Score",
        "Entry", "Stop", "Target", "Stop %", "Target %", "R:R",
        "Size %",
        "RSI", "Vol Ratio", "% from 52WH", "Days Since ER",
        "EPS Gr %", "ROE %", "Forecast 21d %",
        "Top Signals",
    ]
    ws_sw.append(swing_headers)
    _style_header_row(ws_sw)

    swing_candidates = []
    for r in sorted_rs:
        setup = _swing_setup(r)
        if setup is None:
            continue
        swing_candidates.append((r, setup))
    # Sort: conviction tier first, then setup quality
    tier_order = {"A": 0, "B": 1, "C": 2}
    swing_candidates.sort(
        key=lambda x: (tier_order.get(x[1]["conviction"], 9), -x[1]["setup_quality"])
    )

    conv_fill = {
        "A": PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid"),
        "B": PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid"),
        "C": PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
    }

    for rank, (r, sp) in enumerate(swing_candidates, 1):
        f = r.fundamental or {}
        fc = r.forecast or {}
        ws_sw.append([
            rank,
            r.symbol,
            r.market,
            (r.name or "")[:28],
            (r.sector or "")[:20],
            sp["conviction"],
            int(sp["setup_quality"]),
            round(r.composite_score, 1),
            round(r.adjusted_score or r.composite_score, 1),
            sp["entry"],
            sp["stop"],
            sp["target"],
            sp["stop_pct"],
            sp["target_pct"],
            sp["rr"],
            sp["size_pct"],
            round(sp["rsi"], 1),
            round(sp["vol_ratio"], 2),
            round(sp["pct_from_high"], 1),
            sp["er_days"] if sp["er_days"] is not None else "—",
            round((f.get("eps_growth") or 0) * 100, 1),
            round((f.get("roe") or 0) * 100, 1),
            round(fc.get("expected_return_pct") or 0, 2),
            "; ".join(r.all_signals[:4]),
        ])
        row_num = rank + 1
        # Color conviction
        conv_cell = ws_sw.cell(row=row_num, column=swing_headers.index("Conviction") + 1)
        conv_cell.fill = conv_fill.get(sp["conviction"], PatternFill())
        conv_cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        # Color setup quality
        sq_cell = ws_sw.cell(row=row_num, column=swing_headers.index("Setup Q") + 1)
        sq_cell.fill = _score_fill(sp["setup_quality"])
        sq_cell.font = Font(name="Calibri", bold=True, size=10,
                            color="FFFFFF" if sp["setup_quality"] >= 55 or sp["setup_quality"] < 35 else "000000")
        # Color score
        sc_cell = ws_sw.cell(row=row_num, column=swing_headers.index("Score") + 1)
        sc_cell.fill = _score_fill(r.composite_score)
        sc_cell.font = Font(name="Calibri", bold=True, size=10,
                            color="FFFFFF" if r.composite_score >= 55 or r.composite_score < 35 else "000000")
        # Highlight R:R
        rr_cell = ws_sw.cell(row=row_num, column=swing_headers.index("R:R") + 1)
        if sp["rr"] >= 2.5:
            rr_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            rr_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        elif sp["rr"] >= 1.5:
            rr_cell.fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
            rr_cell.font = Font(name="Calibri", bold=True, size=10, color="000000")

    # Border + alignment for all rows
    for row_cells in ws_sw.iter_rows(min_row=2, max_row=ws_sw.max_row, max_col=len(swing_headers)):
        for cell in row_cells:
            cell.border = _BORDER
            if cell.font == Font():
                cell.font = _DATA_FONT
            if cell.column == swing_headers.index("Top Signals") + 1:
                cell.alignment = _LEFT
            else:
                cell.alignment = _CENTER

    # Add a legend / instructions row block above the data via a comment row
    # (kept simple — just freeze + filter)
    ws_sw.freeze_panes = "F2"  # freeze through Conviction column
    ws_sw.auto_filter.ref = f"A1:{get_column_letter(len(swing_headers))}1"
    _auto_width(ws_sw)

    # If no candidates, write a friendly note in row 2
    if not swing_candidates:
        ws_sw.cell(row=2, column=1, value="No swing setups passed filters today.").font = Font(
            name="Calibri", italic=True, size=11, color="808080"
        )
        ws_sw.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(swing_headers))

    # ── Sheet 4: Avoid / Weak ───────────────────────────────────────────
    ws3 = wb.create_sheet(title="Avoid - Weak")
    avoid_headers = ["Ticker", "Name", "Market", "Score", "Verdict", "Price", "RSI", "1M %", "3M %", "Key Reason"]
    ws3.append(avoid_headers)
    _style_header_row(ws3)

    for r in bottom:
        reason = (r.all_signals[0] if r.all_signals else r.error or "weak signals")
        ws3.append([
            r.symbol,
            (r.name or "")[:28],
            r.market,
            round(r.composite_score, 1),
            r.verdict,
            round(r.price, 2),
            round(r.technical.get("rsi", 0), 1),
            round(r.momentum.get("ret_1m") or 0, 2),
            round(r.momentum.get("ret_3m") or 0, 2),
            reason,
        ])

    for row_cells in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(avoid_headers)):
        for cell in row_cells:
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(avoid_headers))}1"
    _auto_width(ws3)

    return wb


def write_reports(reports: list, top_n: int = 15) -> tuple[Path, Path, Path]:
    today = datetime.now().strftime("%Y-%m-%d")
    md = render_markdown(reports, top_n)
    md_path = REPORTS_DIR / f"report_{today}.md"
    md_path.write_text(md, encoding="utf-8")

    json_path = REPORTS_DIR / f"report_{today}.json"
    json_path.write_text(
        json.dumps([r.to_dict() for r in reports], default=str, indent=2),
        encoding="utf-8",
    )

    xlsx_path = REPORTS_DIR / f"report_{today}.xlsx"
    wb = render_excel(reports, top_n=top_n)
    wb.save(str(xlsx_path))

    return md_path, json_path, xlsx_path


def telegram_summary(reports: list, top_n: int = 10) -> str:
    """Compact message for Telegram (4096 char limit)."""
    today = datetime.now().strftime("%Y-%m-%d")
    sorted_rs = sorted(reports, key=lambda r: r.composite_score, reverse=True)
    lines = [f"📊 *Daily Stock Picks — {today}*", f"_{len(reports)} tickers analyzed_\n"]
    lines.append("*🟢 TOP PICKS*")
    for i, r in enumerate(sorted_rs[:top_n], 1):
        flag = "🇮🇳" if r.market == "IN" else "🇺🇸"
        m1 = r.momentum.get("ret_1m")
        m1s = f"{m1:+.1f}%" if m1 is not None else "—"
        sig = (r.all_signals[0] if r.all_signals else "")[:40]
        lines.append(f"{i}. {flag} `{r.symbol}` — *{r.verdict}* ({r.composite_score:.0f}) | 1M {m1s}")
        if sig:
            lines.append(f"   _{sig}_")
    lines.append("\n*🔴 AVOID*")
    for r in sorted_rs[-3:]:
        flag = "🇮🇳" if r.market == "IN" else "🇺🇸"
        lines.append(f"• {flag} `{r.symbol}` — {r.verdict} ({r.composite_score:.0f})")
    lines.append("\n_⚠️ Not investment advice. Full report saved to `reports/`._")
    return "\n".join(lines)
