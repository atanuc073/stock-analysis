"""Walk-forward weight optimizer for the composite score.

Searches for SCORE_WEIGHTS that maximize a chosen objective (Sharpe,
Sortino, Calmar, or CAGR) on out-of-sample data. Two modes:

    1. Single split (default): train on first window, test on second.
    2. Walk-forward (--walk-forward): rolling 3y train / 1y test windows.

Three search strategies:
    - 'random'   — Dirichlet-sampled weight vectors (default; fast, broad).
    - 'sleeves'  — single-factor sleeves (each factor at 1.0, others 0).
                   Used to measure the marginal contribution of each factor;
                   shows you which factors are pulling weight and which are
                   noise.
    - 'grid'     — exhaustive grid over coarse weight steps (e.g. {0, 0.25,
                   0.5, 0.75, 1.0}) across ACTIVE_FACTORS, renormalized to
                   sum=1. Use this when you want sklearn-style GridSearchCV
                   behavior.

Objectives include score-calibration metrics that directly target the
monotonicity problem (Q1..Q5 must increase with score):
    - 'monotonicity' — Spearman rank correlation between score-bucket rank
                       and avg PnL. +1.0 = perfectly monotonic.
    - 'q5_q1'        — Q5 avg PnL minus Q1 avg PnL (top-vs-bottom spread).
    - 'sharpe_mono'  — combined: sharpe + 2*monotonicity (balances risk-
                       adjusted return with rank quality).

Examples:
    # Random search, 100 candidates, single split
    python -m backtest.optimize --start 2019-01-01 --end 2025-01-01 \\
        --candidates 100

    # Diagnose which factors carry the strategy (zero search; one run per factor)
    python -m backtest.optimize --start 2019-01-01 --end 2025-01-01 \\
        --strategy sleeves

    # Walk-forward — 3y train, 1y test, rolling
    python -m backtest.optimize --start 2018-01-01 --end 2025-01-01 \\
        --candidates 80 --walk-forward
"""
from __future__ import annotations
import argparse
import logging
import sys
from dataclasses import asdict
from pathlib import Path

import numpy as np
import pandas as pd

from config import REPORTS_DIR, WATCHLIST, WATCHLIST_INDIA, WATCHLIST_US

from .data_loader import load_universe, trading_dates
from .engine import BacktestConfig, BacktestEngine
from .results import compute as compute_stats, score_calibration

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
log = logging.getLogger("backtest.optimize")

# Components that are scored in backtest live_weights mode (sentiment/options
# are dropped because no historical data; forecast off by default).
ACTIVE_FACTORS = ["technical", "fundamental", "momentum", "quality", "earnings_drift"]


def _resolve_universe(name: str) -> list[str]:
    n = name.lower()
    if n == "watchlist": return WATCHLIST
    if n == "india":     return WATCHLIST_INDIA
    if n == "us":        return WATCHLIST_US
    if n == "nifty500":
        try:
            from data_sources.universe import nifty500_tickers
            return nifty500_tickers()
        except Exception as e:
            log.warning("nifty500 unavailable (%s); fallback to watchlist", e)
            return WATCHLIST
    if n == "nse_all":
        try:
            from data_sources.universe import nse_all_tickers
            return nse_all_tickers()
        except Exception as e:
            log.warning("nse_all unavailable (%s); fallback to watchlist", e)
            return WATCHLIST
    return [s.strip() for s in name.split(",") if s.strip()]


def _dirichlet_weights(n_candidates: int, factors: list[str],
                       seed: int = 42) -> list[dict]:
    """Sample weight vectors uniformly from the simplex (sum=1, each >=0).

    Uses Dirichlet(1, 1, ..., 1) which is uniform on the simplex.
    """
    rng = np.random.default_rng(seed)
    samples = rng.dirichlet(np.ones(len(factors)), size=n_candidates)
    out = []
    for row in samples:
        w = {f: float(round(row[i], 4)) for i, f in enumerate(factors)}
        # Add zero-weighted slots for non-active factors so config-shape
        # is preserved end-to-end.
        for k in ("sentiment", "options", "forecast", "valuation"):
            w.setdefault(k, 0.0)
        # Renormalize after rounding drift
        total = sum(w[f] for f in factors)
        if total > 0:
            for f in factors:
                w[f] = w[f] / total
        out.append(w)
    return out


def _sleeve_weights(factors: list[str]) -> list[dict]:
    """One weight vector per factor: that factor at 1.0, all others at 0."""
    out = []
    for active in factors:
        w = {f: (1.0 if f == active else 0.0) for f in factors}
        for k in ("sentiment", "options", "forecast", "valuation"):
            w.setdefault(k, 0.0)
        out.append(w)
    return out


def _grid_weights(factors: list[str], steps: list[float],
                  min_active: int = 2) -> list[dict]:
    """Exhaustive grid over ``steps`` per factor; renormalize sum→1.

    Drops candidates with all-zero weights or fewer than ``min_active`` non-
    zero factors (to avoid degenerate single-factor + noise combinations
    that the 'sleeves' strategy already covers).

    Combinatorial size = len(steps) ** len(factors). With 5 factors and
    5 steps that's 3,125 candidates → use 3 or 4 steps for practical runs.
    """
    import itertools
    out: list[dict] = []
    seen: set[tuple] = set()
    for combo in itertools.product(steps, repeat=len(factors)):
        if sum(combo) <= 0:
            continue
        if sum(1 for v in combo if v > 0) < min_active:
            continue
        total = sum(combo)
        norm = tuple(round(v / total, 4) for v in combo)
        if norm in seen:
            continue
        seen.add(norm)
        w = {f: float(norm[i]) for i, f in enumerate(factors)}
        for k in ("sentiment", "options", "forecast", "valuation"):
            w.setdefault(k, 0.0)
        out.append(w)
    return out


def _calibration_metrics(result) -> tuple[float, float]:
    """Compute (monotonicity, q5_q1_spread) from a backtest result.

    monotonicity = Spearman rank corr between bucket rank (Q1=1..Q5=5) and
                   AvgPnL_Pct. +1.0 = perfectly monotonic, -1.0 = perfectly
                   inverted (the current Q5<Q1 problem).
    q5_q1_spread = AvgPnL_Pct[Q5] - AvgPnL_Pct[Q1] (percentage points).
    """
    try:
        calib = score_calibration(result, n_quantiles=5)
        if calib.empty or len(calib) < 2:
            return 0.0, 0.0
        # Bucket labels look like "Q1 [70.15-76.17]" → extract Q index
        calib = calib.copy()
        calib["q_idx"] = calib["Score_Bucket"].str.extract(r"Q(\d+)").astype(int)
        calib = calib.sort_values("q_idx")
        # Spearman corr between rank order and avg PnL
        mono = float(calib["q_idx"].corr(calib["AvgPnL_Pct"], method="spearman"))
        if not np.isfinite(mono):
            mono = 0.0
        spread = float(calib["AvgPnL_Pct"].iloc[-1] - calib["AvgPnL_Pct"].iloc[0])
        return mono, spread
    except Exception as e:
        log.debug("calibration metrics failed: %s", e)
        return 0.0, 0.0


def _run_backtest(data, dates, weights: dict, capital: float,
                  threshold: float, max_pos: int) -> dict:
    """Run one backtest with given weights; return key stats dict."""
    cfg = BacktestConfig(
        initial_capital=capital,
        rebalance_freq_days=5,
        min_score=threshold,
        max_positions=max_pos,
        include_forecast=False,
        live_weights=True,
        use_regime=True,
        weights=weights,
    )
    engine = BacktestEngine(data, cfg)
    result = engine.run(dates)
    fail = {"sharpe": -99.0, "cagr": -99.0, "calmar": -99.0,
            "sortino": -99.0, "trades": 0, "max_dd": 0.0, "final": capital,
            "win_rate": 0.0, "monotonicity": -1.0, "q5_q1": -99.0}
    if not result.equity_curve:
        return fail
    try:
        stats = compute_stats(result)
    except Exception as e:
        log.warning("stats failed: %s", e)
        return fail
    mono, spread = _calibration_metrics(result)
    return {
        "sharpe":       float(stats.sharpe_ratio),
        "sortino":      float(stats.sortino_ratio),
        "calmar":       float(stats.calmar_ratio),
        "cagr":         float(stats.cagr_pct),
        "max_dd":       float(stats.max_drawdown_pct),
        "trades":       int(stats.total_trades),
        "win_rate":     float(stats.win_rate_pct),
        "final":        float(stats.final_equity),
        "monotonicity": float(mono),
        "q5_q1":        float(spread),
    }


def _objective_value(stats: dict, objective: str) -> float:
    """Map a stats dict to a single scalar for ranking."""
    if stats["trades"] < 5:
        return -99.0  # exclude pathological zero-trade runs
    if objective == "sharpe":       return stats["sharpe"]
    if objective == "sortino":      return stats["sortino"]
    if objective == "calmar":       return stats["calmar"]
    if objective == "cagr":         return stats["cagr"]
    if objective == "monotonicity": return stats["monotonicity"]
    if objective == "q5_q1":        return stats["q5_q1"]
    # Combined: rewards both risk-adjusted return AND a monotonic score curve.
    # Sharpe is ~[-1, 3], monotonicity is [-1, 1], so 2x weight keeps scales
    # comparable. Tune the 2.0 if you want to favor one side more.
    if objective == "sharpe_mono":
        return stats["sharpe"] + 2.0 * stats["monotonicity"]
    raise ValueError(f"unknown objective: {objective}")


def _slice_dates(dates: pd.DatetimeIndex, start: str, end: str) -> pd.DatetimeIndex:
    s = pd.Timestamp(start); e = pd.Timestamp(end)
    return dates[(dates >= s) & (dates <= e)]


def _walk_forward_windows(start: str, end: str,
                           train_years: int = 3,
                           test_years: int = 1) -> list[tuple[str, str, str, str]]:
    """Generate (train_start, train_end, test_start, test_end) tuples."""
    s = pd.Timestamp(start); e = pd.Timestamp(end)
    windows = []
    cur = s
    while cur + pd.DateOffset(years=train_years + test_years) <= e:
        ts = cur
        te = ts + pd.DateOffset(years=train_years)
        vs = te + pd.Timedelta(days=1)
        ve = vs + pd.DateOffset(years=test_years)
        windows.append((ts.strftime("%Y-%m-%d"), te.strftime("%Y-%m-%d"),
                        vs.strftime("%Y-%m-%d"), ve.strftime("%Y-%m-%d")))
        cur = cur + pd.DateOffset(years=test_years)
    return windows


def main(argv: list[str] | None = None) -> int:
    def _safe_to_csv(df_to_save: pd.DataFrame, path: Path):
        try:
            df_to_save.to_csv(path, index=False)
            log.info("Successfully saved results to %s", path)
        except PermissionError:
            import time
            timestamp = int(time.time())
            backup_path = path.parent / f"{path.stem}_backup_{timestamp}.csv"
            log.error("❌ Permission Denied when saving to %s! The file might be open in Excel.", path)
            log.warning("⚠️ Saving to backup file instead: %s", backup_path)
            try:
                df_to_save.to_csv(backup_path, index=False)
                log.info("Successfully saved backup to %s", backup_path)
            except Exception as e:
                log.critical("🚨 Failed to save even the backup file: %s", e)
        except Exception as e:
            log.critical("🚨 Unexpected error saving to %s: %s", path, e)

    def _save_excel(path: Path, df: pd.DataFrame, fold_top10s: dict[int, pd.DataFrame], avg_weights: dict[str, float], objective: str):
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as w:
                df.to_excel(w, sheet_name="Walk_Forward_Summary", index=False)
                
                weight_rows = [{"Factor": k, "Weight": v} for k, v in avg_weights.items()]
                pd.DataFrame(weight_rows).to_excel(w, sheet_name="Suggested_Weights", index=False)
                
                for f_num, f_df in fold_top10s.items():
                    f_df.to_excel(w, sheet_name=f"Fold{f_num}_Top10_Candidates", index=False)
            
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = load_workbook(path)
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(bold=True, color="FFFFFF")
            
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                ws.row_dimensions[1].height = 22
                ws.freeze_panes = "B2"
                for col in ws.columns:
                    try:
                        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                    except ValueError:
                        max_len = 12
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)
                if ws.max_row > 1:
                    ws.auto_filter.ref = ws.dimensions
            wb.save(path)
            log.info("✨ Successfully saved beautiful Excel optimization report to %s", path)
        except PermissionError:
            import time
            timestamp = int(time.time())
            backup_path = path.parent / f"{path.stem}_backup_{timestamp}.xlsx"
            log.error("❌ Permission Denied when saving Excel report to %s! The file might be open in Excel.", path)
            log.warning("⚠️ Saving Excel report to backup instead: %s", backup_path)
            try:
                _save_excel(backup_path, df, fold_top10s, avg_weights, objective)
            except Exception as e:
                log.critical("🚨 Failed to save Excel backup file: %s", e)
        except Exception as e:
            log.error("⚠️ Could not write Excel optimization report: %s", e)

    p = argparse.ArgumentParser(description="Walk-forward weight optimizer.")
    p.add_argument("--start", required=True)
    p.add_argument("--end", required=True)
    p.add_argument("--universe", default="watchlist")
    p.add_argument("--capital", type=float, default=1_000_000.0)
    p.add_argument("--threshold", type=float, default=70.0)
    p.add_argument("--max-positions", type=int, default=12)
    p.add_argument("--candidates", type=int, default=80,
                   help="random search candidates (ignored for sleeves/grid)")
    p.add_argument("--strategy", choices=["random", "sleeves", "grid"],
                   default="random")
    p.add_argument("--grid-steps", default="0,0.25,0.5,0.75,1.0",
                   help="comma-separated weight steps for --strategy grid")
    p.add_argument("--grid-min-active", type=int, default=2,
                   help="min number of non-zero factors per grid candidate")
    p.add_argument("--objective",
                   choices=["sharpe", "sortino", "calmar", "cagr",
                            "monotonicity", "q5_q1", "sharpe_mono"],
                   default="sharpe")
    p.add_argument("--rank-by",
                   choices=["train", "test", "consistent", "min"],
                   default="train",
                   help="how to pick per-fold winner: 'train' (legacy, "
                        "overfits), 'test' (OOS-best), 'consistent' (test "
                        "minus 0.5*|train-test|; rewards small generalization "
                        "gap — RECOMMENDED), 'min' (worst-of-two; most "
                        "conservative). Anything other than 'train' costs "
                        "2x backtests since every candidate is tested.")
    p.add_argument("--gap-penalty", type=float, default=0.5,
                   help="lambda in 'consistent' rank-by: test - lambda*|train-test|")
    p.add_argument("--walk-forward", action="store_true",
                   help="rolling 3y train / 1y test windows")
    p.add_argument("--train-years", type=int, default=3)
    p.add_argument("--test-years", type=int, default=1)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--max-workers", type=int, default=4)
    p.add_argument("--sample-size", type=int, default=None,
                   help="randomly sample N symbols from the resolved universe")
    p.add_argument("--output-dir", default=None)
    args = p.parse_args(argv)

    output_dir = Path(args.output_dir) if args.output_dir else REPORTS_DIR / "optimize"
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Generate candidate weights ───────────────────────────────────
    if args.strategy == "sleeves":
        candidates = _sleeve_weights(ACTIVE_FACTORS)
        log.info("Strategy=sleeves: testing %d single-factor configurations",
                 len(candidates))
    elif args.strategy == "grid":
        steps = [float(s) for s in args.grid_steps.split(",") if s.strip()]
        candidates = _grid_weights(ACTIVE_FACTORS, steps,
                                   min_active=args.grid_min_active)
        log.info("Strategy=grid: %d candidates from steps=%s (min_active=%d)",
                 len(candidates), steps, args.grid_min_active)
    else:
        candidates = _dirichlet_weights(args.candidates, ACTIVE_FACTORS, seed=args.seed)
        log.info("Strategy=random: %d Dirichlet-sampled candidates", len(candidates))

    # ── Load data once for the full span ─────────────────────────────
    symbols = _resolve_universe(args.universe)
    if args.sample_size and len(symbols) > args.sample_size:
        import random
        random.seed(args.seed)
        symbols = random.sample(symbols, args.sample_size)
        log.info("Randomly sampled %d symbols out of %s (seed=%d)", len(symbols), args.universe, args.seed)

    data = load_universe(symbols, args.start, args.end, max_workers=args.max_workers)
    if not data:
        log.error("No data loaded; aborting")
        return 1
    log.info("Loaded %d/%d symbols", len(data), len(symbols))

    all_dates = trading_dates(data, args.start, args.end)
    if len(all_dates) == 0:
        log.error("No trading dates")
        return 1

    # ── Single split or walk-forward ─────────────────────────────────
    if args.walk_forward:
        windows = _walk_forward_windows(args.start, args.end,
                                         args.train_years, args.test_years)
        if not windows:
            log.error("Span too short for %dy/%dy walk-forward",
                      args.train_years, args.test_years)
            return 1
        log.info("Walk-forward: %d windows", len(windows))
    else:
        # Use 70%/30% split
        split_idx = int(len(all_dates) * 0.7)
        ts = all_dates[0].strftime("%Y-%m-%d")
        te = all_dates[split_idx].strftime("%Y-%m-%d")
        vs = all_dates[split_idx + 1].strftime("%Y-%m-%d") if split_idx + 1 < len(all_dates) else te
        ve = all_dates[-1].strftime("%Y-%m-%d")
        windows = [(ts, te, vs, ve)]
        log.info("Single split: train=[%s..%s] test=[%s..%s]", ts, te, vs, ve)

    # ── For each window, score every candidate on train; record OOS ──
    fold_records: list[dict] = []
    fold_top10s: dict[int, pd.DataFrame] = {}
    for fold_idx, (ts, te, vs, ve) in enumerate(windows):
        train_dates = _slice_dates(all_dates, ts, te)
        test_dates = _slice_dates(all_dates, vs, ve)
        log.info("Fold %d: train %s..%s (%d days), test %s..%s (%d days)",
                 fold_idx + 1, ts, te, len(train_dates), vs, ve, len(test_dates))

        # 1) score every candidate on the TRAIN window
        train_results = []
        best_so_far_score = -999.0
        for ci, w in enumerate(candidates):
            stats = _run_backtest(data, train_dates, w,
                                  args.capital, args.threshold, args.max_positions)
            score = _objective_value(stats, args.objective)
            train_results.append((ci, score, stats))
            
            # Format active weights nicely
            w_str = ", ".join(f"{f[:4].capitalize()}:{w[f]:.2f}" for f in ACTIVE_FACTORS)
            
            # Check if this candidate is the new best
            is_new_best = score > best_so_far_score
            if is_new_best:
                best_so_far_score = score
                best_indicator = " ✨ [NEW BEST]"
            else:
                best_indicator = ""
                
            log.info("  Train Candidate %d/%d [%s] -> Score: %.3f (Sharpe: %.2f, Mono: %.2f, CAGR: %.1f%%, DD: %.1f%%)%s",
                     ci + 1, len(candidates), w_str, score, stats["sharpe"], stats["monotonicity"], stats["cagr"], stats["max_dd"], best_indicator)

        # 2) Decide who wins this fold.
        # Legacy behavior: pick the candidate with the highest TRAIN score and
        # test only that one. Cheap but the textbook overfitting trap — the
        # best-train candidate often collapses OOS (e.g. train mono +0.9 →
        # test mono -0.2). The newer rank-by modes test EVERY candidate on
        # the test window too, then rank by a metric that explicitly rewards
        # train→test stability.
        per_candidate_test: dict[int, dict] = {}
        if args.rank_by == "train":
            # Cheap path — only test the train winner.
            train_results.sort(key=lambda x: x[1], reverse=True)
            best_ci, best_train_score, best_train_stats = train_results[0]
            best_w = candidates[best_ci]
            test_stats = _run_backtest(data, test_dates, best_w,
                                       args.capital, args.threshold, args.max_positions)
            best_test_score = _objective_value(test_stats, args.objective)
            per_candidate_test[best_ci] = test_stats
        else:
            # Honest path — score every candidate on the test window too.
            log.info("  test: scoring all %d candidates OOS (rank-by=%s)",
                     len(candidates), args.rank_by)
            for ci, _, _ in train_results:
                ts_stats = _run_backtest(data, test_dates, candidates[ci],
                                         args.capital, args.threshold,
                                         args.max_positions)
                per_candidate_test[ci] = ts_stats
                
                # Format active weights nicely
                w = candidates[ci]
                w_str = ", ".join(f"{f[:4].capitalize()}:{w[f]:.2f}" for f in ACTIVE_FACTORS)
                te_score = _objective_value(ts_stats, args.objective)
                
                log.info("  Test Candidate %d/%d [%s] -> Score: %.3f (Sharpe: %.2f, Mono: %.2f, CAGR: %.1f%%, DD: %.1f%%)",
                         ci + 1, len(candidates), w_str, te_score, ts_stats["sharpe"], ts_stats["monotonicity"], ts_stats["cagr"], ts_stats["max_dd"])

            # Combined metric per candidate
            def _combined(ci: int, tr_score: float) -> float:
                te_score = _objective_value(per_candidate_test[ci], args.objective)
                if not np.isfinite(te_score) or te_score <= -98:
                    return -99.0
                if args.rank_by == "test":
                    return te_score
                if args.rank_by == "min":
                    return min(tr_score, te_score)
                # 'consistent': reward small train→test gap
                return te_score - args.gap_penalty * abs(tr_score - te_score)

            scored = [(ci, _combined(ci, tr_s), tr_s, stats)
                      for ci, tr_s, stats in train_results]
            scored.sort(key=lambda x: x[1], reverse=True)
            best_ci, _, best_train_score, best_train_stats = scored[0]
            best_w = candidates[best_ci]
            test_stats = per_candidate_test[best_ci]
            best_test_score = _objective_value(test_stats, args.objective)
            # Replace train_results ordering so the top-10 CSV reflects the
            # chosen rank-by criterion, not raw train score.
            train_results = [(ci, tr_s, stats) for ci, _, tr_s, stats in scored]

        test_score = best_test_score

        fold_records.append({
            "fold": fold_idx + 1, "train": f"{ts}..{te}", "test": f"{vs}..{ve}",
            "best_candidate": best_ci,
            "rank_by": args.rank_by,
            "train_score": round(best_train_score, 3),
            "test_score": round(test_score, 3),
            "gap": round(abs(best_train_score - test_score), 3),
            "test_cagr": round(test_stats["cagr"], 2),
            "test_sharpe": round(test_stats["sharpe"], 3),
            "test_max_dd": round(test_stats["max_dd"], 2),
            "test_trades": test_stats["trades"],
            "test_mono": round(test_stats["monotonicity"], 3),
            "test_q5_q1": round(test_stats["q5_q1"], 2),
            **{f"w_{k}": round(v, 3) for k, v in best_w.items()
               if k in ACTIVE_FACTORS},
        })

        # Also keep top-10 by-train table for diagnostics. When rank-by != train,
        # the order reflects the combined criterion (test-honest) instead.
        top10 = []
        for ci, s, st in train_results[:10]:
            te_stats = per_candidate_test.get(ci)
            te_score = (_objective_value(te_stats, args.objective)
                        if te_stats else None)
            row = {"candidate": ci, "train_score": round(s, 3),
                   "test_score": round(te_score, 3) if te_score is not None else "",
                   "gap": round(abs(s - te_score), 3) if te_score is not None else "",
                   "trades": st["trades"], "cagr": round(st["cagr"], 2),
                   "sharpe": round(st["sharpe"], 3),
                   "max_dd": round(st["max_dd"], 2),
                   "mono": round(st["monotonicity"], 3),
                   "q5_q1": round(st["q5_q1"], 2),
                   "test_mono": (round(te_stats["monotonicity"], 3)
                                 if te_stats else ""),
                   "test_q5_q1": (round(te_stats["q5_q1"], 2)
                                  if te_stats else "")}
            row.update({f"w_{k}": round(candidates[ci][k], 3)
                        for k in ACTIVE_FACTORS})
            top10.append(row)
        top10_df = pd.DataFrame(top10)
        _safe_to_csv(top10_df, output_dir / f"fold{fold_idx + 1}_top10_train.csv")
        fold_top10s[fold_idx + 1] = top10_df

    # ── Save & summarize ─────────────────────────────────────────────
    df = pd.DataFrame(fold_records)
    out_csv = output_dir / "walk_forward_results.csv"
    _safe_to_csv(df, out_csv)

    print("\n" + "=" * 80)
    print(f"WALK-FORWARD RESULTS (objective: {args.objective})")
    print("=" * 80)
    print(df.to_string(index=False))
    print()

    # Average best-weights across folds (the most defensible "deploy" choice)
    avg_weights = {}
    if len(df) > 0:
        avg_weights = {f: float(df[f"w_{f}"].mean()) for f in ACTIVE_FACTORS}
        total = sum(avg_weights.values())
        if total > 0:
            avg_weights = {k: round(v / total, 3) for k, v in avg_weights.items()}
        print("Mean OOS test score:   ", round(df["test_score"].mean(), 3))
        print("Median OOS test score: ", round(df["test_score"].median(), 3))
        print()
        print("Suggested SCORE_WEIGHTS (mean of per-fold winners, renormalized):")
        for k, v in avg_weights.items():
            print(f"    \"{k}\":\t{v:.3f},")
        print()
        
    # Save the beautiful Excel workbook report!
    _save_excel(output_dir / "optimization_results.xlsx", df, fold_top10s, avg_weights, args.objective)
    
    print(f"Detailed results: {out_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())