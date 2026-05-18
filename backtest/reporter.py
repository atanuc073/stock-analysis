"""Backtest reporter — Excel workbook, Markdown summary, equity-curve PNG."""
from __future__ import annotations
import logging
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

from .engine import BacktestResult
from .results import (
    PerformanceStats,
    yearly_breakdown,
    benchmark_buy_and_hold,
    benchmark_sip,
    score_calibration,
    top_chase_diagnostic,
    regime_breakdown,
)

log = logging.getLogger(__name__)


def write_excel(result: BacktestResult, stats: PerformanceStats,
                benchmarks: dict[str, pd.Series], path: Path,
                data: Optional[dict] = None) -> None:
    """Multi-sheet Excel with summary, trades, equity curve, yearly P&L."""
    eq_df = pd.DataFrame([
        {"Date": p.date, "Equity": p.total, "Cash": p.cash,
         "MarketValue": p.market_value, "OpenPositions": p.n_open,
         "Regime": p.regime.label if hasattr(p, "regime") and p.regime else ""}
        for p in result.equity_curve
    ])

    trades_df = pd.DataFrame([
        {"Date": t.timestamp[:10], "Symbol": t.symbol, "Mkt": t.market,
         "Sector": t.sector, "Action": t.action, "Qty": round(t.qty, 4),
         "Price": round(t.price, 2), "Gross": round(t.gross_value, 0),
         "Cost": round(t.cost, 2), "Net": round(t.net_value, 0),
         "PnL_Abs": round(t.pnl_abs, 0), "PnL_%": round(t.pnl_pct, 2),
         "Days_Held": t.days_held, "Exit_Type": t.exit_type or "",
         "Score_Entry": round(t.uptrend_score_at_entry if result.config.uptrend_mode else t.score_at_entry, 1), "Reason": t.reason}
        for t in result.trades
    ])

    sip_active = bool(result.contributions) and result.config.sip_amount > 0
    summary_rows = [
        ("Period", f"{result.start} → {result.end}"),
        ("Years", f"{stats.years:.2f}"),
        ("Mode", "SIP" if sip_active else "Lumpsum"),
        ("Initial Capital", f"₹{stats.initial_capital:,.0f}"),
    ]
    if sip_active:
        summary_rows += [
            ("SIP Amount/Month", f"₹{result.config.sip_amount:,.0f}"),
            ("SIP Day", result.config.sip_day_of_month),
            ("Total Contributions", stats.n_contributions),
            ("Total Invested", f"₹{stats.total_invested:,.0f}"),
        ]
    summary_rows += [
        ("Final Equity", f"₹{stats.final_equity:,.0f}"),
        ("Total Return", f"{stats.total_return_pct:+.2f}%"),
        ("CAGR" if not sip_active else "XIRR (annualized)",
         f"{stats.cagr_pct:+.2f}%"),
        ("", ""),
        ("RISK", ""),
        ("Max Drawdown", f"{stats.max_drawdown_pct:.2f}%"),
        ("DD Duration (days)", stats.max_drawdown_duration_days),
        ("Annualized Volatility", f"{stats.annual_volatility_pct:.2f}%"),
        ("Sharpe Ratio", f"{stats.sharpe_ratio:.2f}"),
        ("Sortino Ratio", f"{stats.sortino_ratio:.2f}"),
        ("Calmar Ratio", f"{stats.calmar_ratio:.2f}"),
        ("", ""),
        ("TRADES", ""),
        ("Total Trades", stats.total_trades),
        ("Closed (Sells)", stats.closed_trades),
        ("Win Rate", f"{stats.win_rate_pct:.1f}%"),
        ("Avg Win", f"{stats.avg_win_pct:+.2f}%"),
        ("Avg Loss", f"{stats.avg_loss_pct:+.2f}%"),
        ("Expectancy/Trade", f"{stats.expectancy_pct:+.2f}%"),
        ("Profit Factor", f"{stats.profit_factor:.2f}"),
        ("Avg Hold (days)", f"{stats.avg_holding_days:.0f}"),
        ("", ""),
        ("REGIMES (Market Context)", ""),
    ]
    
    # Count days in each regime
    regime_counts = {}
    for p in result.equity_curve:
        label = p.regime.label if hasattr(p, "regime") and p.regime else "Unknown"
        regime_counts[label] = regime_counts.get(label, 0) + 1
        
    for label, count in sorted(regime_counts.items(), key=lambda x: x[1], reverse=True):
        summary_rows.append((f"  Days in {label}", count))

    for k, v in stats.exits_by_type.items():
        summary_rows.append((f"  Exit: {k}", v))
    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    yearly_df = yearly_breakdown(result)

    # Per-symbol P&L summary
    if not trades_df.empty:
        sells = trades_df[trades_df["Action"] == "SELL"]
        per_sym = (sells.groupby(["Symbol", "Mkt", "Sector"])
                       .agg(Trades=("PnL_Abs", "count"),
                            TotalPnL=("PnL_Abs", "sum"),
                            AvgPnL_Pct=("PnL_%", "mean"),
                            WinRate=("PnL_%", lambda s: (s > 0).mean() * 100))
                       .round(2).reset_index()
                       .sort_values("TotalPnL", ascending=False))
    else:
        per_sym = pd.DataFrame()

    # Exit type analysis
    if not trades_df.empty:
        sells = trades_df[trades_df["Action"] == "SELL"]
        exit_analysis = (sells.groupby("Exit_Type")
                              .agg(Count=("PnL_Abs", "count"),
                                   AvgPnL_Pct=("PnL_%", "mean"),
                                   TotalPnL=("PnL_Abs", "sum"),
                                   AvgDays=("Days_Held", "mean"))
                              .round(2).reset_index())
    else:
        exit_analysis = pd.DataFrame()

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        summary_df.to_excel(w, sheet_name="Summary", index=False)
        yearly_df.to_excel(w, sheet_name="Yearly", index=False)
        eq_df.to_excel(w, sheet_name="Equity_Curve", index=False)
        if not per_sym.empty:
            per_sym.to_excel(w, sheet_name="By_Symbol", index=False)
        if not exit_analysis.empty:
            exit_analysis.to_excel(w, sheet_name="By_Exit_Type", index=False)
        calib = score_calibration(result)
        if not calib.empty:
            calib.to_excel(w, sheet_name="Score_Calibration", index=False)
        regime_df = regime_breakdown(result)
        if not regime_df.empty:
            regime_df.to_excel(w, sheet_name="By_Regime", index=False)
        # Top-chase diagnostic (only if historical data is provided)
        if data is not None:
            tc = top_chase_diagnostic(result, data)
            if not tc["trades"].empty:
                tc["trades"].to_excel(w, sheet_name="TopChase_Trades", index=False)
            if not tc["buckets_extension"].empty:
                tc["buckets_extension"].to_excel(w, sheet_name="TopChase_By_Extension", index=False)
            if not tc["buckets_52w"].empty:
                tc["buckets_52w"].to_excel(w, sheet_name="TopChase_By_52wHigh", index=False)
            if tc["summary"]:
                summ_df = pd.DataFrame(
                    [(k, v) for k, v in tc["summary"].items()],
                    columns=["Metric", "Value"],
                )
                summ_df.to_excel(w, sheet_name="TopChase_Summary", index=False)
        if not trades_df.empty:
            trades_df.to_excel(w, sheet_name="All_Trades", index=False)

    _format_excel(path)


def _format_excel(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "B2"
        for col in ws.columns:
            try:
                max_len = max(len(str(c.value)) if c.value is not None else 0
                              for c in col)
            except ValueError:
                max_len = 12
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_markdown(result: BacktestResult, stats: PerformanceStats, path: Path,
                    data: Optional[dict] = None) -> None:
    L: list[str] = []
    L.append(f"# Backtest Report — {result.start} → {result.end}\n")
    L.append(f"**Universe:** {result.universe_size} symbols  ")
    L.append(f"**Rebalance frequency:** every {result.config.rebalance_freq_days} days  ")
    L.append(f"**Min score to buy:** {result.config.min_score}  ")
    L.append(f"**Max positions:** {result.config.max_positions}  ")
    L.append(f"**Transaction costs:** {result.config.transaction_cost_bps + result.config.slippage_bps:.0f} bps round-trip\n")

    L.append("## 📈 Performance\n")
    sip_active = bool(result.contributions) and result.config.sip_amount > 0
    L.append("| Metric | Value |")
    L.append("|---|---|")
    L.append(f"| Mode | {'SIP' if sip_active else 'Lumpsum'} |")
    L.append(f"| Initial Capital | ₹{stats.initial_capital:,.0f} |")
    if sip_active:
        L.append(f"| SIP Amount/Month | ₹{result.config.sip_amount:,.0f} (day {result.config.sip_day_of_month}) |")
        L.append(f"| Total Contributions | {stats.n_contributions} |")
        L.append(f"| **Total Invested** | **₹{stats.total_invested:,.0f}** |")
    L.append(f"| Final Equity | ₹{stats.final_equity:,.0f} |")
    L.append(f"| **Total Return** | **{stats.total_return_pct:+.2f}%** |")
    if sip_active:
        L.append(f"| **XIRR (annualized)** | **{stats.xirr_pct:+.2f}%** |")
    else:
        L.append(f"| **CAGR** | **{stats.cagr_pct:+.2f}%** |")
    L.append(f"| Years | {stats.years:.2f} |")

    L.append("\n## ⚠️ Risk\n")
    L.append("| Metric | Value |")
    L.append("|---|---|")
    L.append(f"| **Max Drawdown** | **{stats.max_drawdown_pct:.2f}%** |")
    L.append(f"| Drawdown Duration | {stats.max_drawdown_duration_days} days |")
    L.append(f"| Annualized Volatility | {stats.annual_volatility_pct:.2f}% |")
    L.append(f"| Sharpe Ratio | {stats.sharpe_ratio:.2f} |")
    L.append(f"| Sortino Ratio | {stats.sortino_ratio:.2f} |")
    L.append(f"| Calmar Ratio | {stats.calmar_ratio:.2f} |")

    L.append("\n## 📊 Trade Statistics\n")
    L.append("| Metric | Value |")
    L.append("|---|---|")
    L.append(f"| Total Trades | {stats.total_trades} |")
    L.append(f"| Closed Sells | {stats.closed_trades} |")
    L.append(f"| **Win Rate** | **{stats.win_rate_pct:.1f}%** |")
    L.append(f"| Avg Win | {stats.avg_win_pct:+.2f}% |")
    L.append(f"| Avg Loss | {stats.avg_loss_pct:+.2f}% |")
    L.append(f"| **Expectancy/Trade** | **{stats.expectancy_pct:+.2f}%** |")
    L.append(f"| Profit Factor | {stats.profit_factor:.2f} |")
    L.append(f"| Avg Holding Days | {stats.avg_holding_days:.0f} |")

    L.append("\n### 🌐 Market Context (Regime Breakdown)\n")
    L.append("| Regime | Days | % of Period |")
    L.append("|---|---|---|")
    total_days = len(result.equity_curve)
    regime_counts = {}
    for p in result.equity_curve:
        label = p.regime.label if hasattr(p, "regime") and p.regime else "Unknown"
        regime_counts[label] = regime_counts.get(label, 0) + 1
    
    for label, count in sorted(regime_counts.items(), key=lambda x: x[1], reverse=True):
        pct = (count / total_days) * 100 if total_days > 0 else 0
        L.append(f"| {label} | {count} | {pct:.1f}% |")

    L.append("\n### Exit Reason Breakdown\n")
    L.append("| Exit Type | Count |")
    L.append("|---|---|")
    for exit_type, count in sorted(stats.exits_by_type.items(),
                                   key=lambda x: x[1], reverse=True):
        L.append(f"| {exit_type} | {count} |")

    # Yearly breakdown
    yearly = yearly_breakdown(result)
    if not yearly.empty:
        L.append("\n## 📅 Yearly Performance\n")
        L.append(yearly.to_markdown(index=False))

    # Best & worst trades
    sells = [t for t in result.trades if t.action == "SELL"]
    if sells:
        sells_sorted = sorted(sells, key=lambda t: t.pnl_abs, reverse=True)
        L.append("\n## 🏆 Top 10 Winners\n")
        L.append("| Symbol | Mkt | Entry → Exit | P&L | Days | Reason |")
        L.append("|---|---|---|---|---|---|")
        for t in sells_sorted[:10]:
            L.append(f"| `{t.symbol}` | {t.market} | {t.timestamp[:10]} | "
                     f"₹{t.pnl_abs:+,.0f} ({t.pnl_pct:+.1f}%) | {t.days_held} | {t.exit_type} |")

        L.append("\n## 💔 Worst 10 Trades\n")
        L.append("| Symbol | Mkt | Date | P&L | Days | Reason |")
        L.append("|---|---|---|---|---|---|")
        for t in sells_sorted[-10:]:
            L.append(f"| `{t.symbol}` | {t.market} | {t.timestamp[:10]} | "
                     f"₹{t.pnl_abs:+,.0f} ({t.pnl_pct:+.1f}%) | {t.days_held} | {t.exit_type} |")

    # Score calibration — does score predict forward returns?
    calib = score_calibration(result)
    if not calib.empty:
        L.append("\n## 🎯 Score Calibration (does the scorer actually predict returns?)\n")
        L.append("Buckets are **quantile-based** (each bucket holds ~equal trades). "
                 "If `AvgPnL_Pct` / `WinRate_Pct` rise monotonically with bucket, the "
                 "scorer ranks trades correctly. Flat = noise above the threshold.\n")
        L.append(calib.to_markdown(index=False))

    # Per-regime P&L — are returns concentrated in one regime?
    regime_df = regime_breakdown(result)
    if not regime_df.empty:
        L.append("\n## 🌡️ P&L by Entry Regime\n")
        L.append("_Profit concentrated in BULL only = regime-conditional bet. "
                 "Steady across regimes = genuine all-weather alpha._\n")
        L.append(regime_df.to_markdown(index=False))

    # Top-chase diagnostic — are we buying near tops?
    if data is not None:
        tc = top_chase_diagnostic(result, data)
        summ = tc.get("summary") or {}
        if summ:
            L.append("\n## 🔍 Top-Chase Diagnostic (are we buying near tops?)\n")
            L.append("Empirical check: for every BUY, what was the entry context, "
                     "and what did the stock do over the next 30/90 days?\n")
            L.append("### Summary\n")
            L.append("| Indicator | Value |")
            L.append("|---|---|")
            for k, v in summ.items():
                L.append(f"| {k} | {v} |")
            if not tc["buckets_extension"].empty:
                L.append("\n### Forward returns by extension above 200DMA\n")
                L.append("_If `Avg_Fwd30` declines as extension grows, you're chasing tops._\n")
                L.append(tc["buckets_extension"].to_markdown(index=False))
            if not tc["buckets_52w"].empty:
                L.append("\n### Forward returns by distance from 52-week high\n")
                L.append("_Pullback buckets should outperform near-high buckets._\n")
                L.append(tc["buckets_52w"].to_markdown(index=False))

    # Verdict
    L.append("\n## 🎯 Verdict\n")
    if stats.cagr_pct >= 15 and stats.max_drawdown_pct >= -25 and stats.sharpe_ratio >= 1.0:
        L.append("✅ **Strategy passes**. CAGR > 15%, DD < 25%, Sharpe > 1.0.")
    elif stats.cagr_pct >= 10:
        L.append("⚠️ **Strategy works but margins are thin.** Tune thresholds and re-run.")
    else:
        L.append("❌ **Strategy fails baseline.** CAGR too low or DD too deep. Do NOT deploy capital.")

    L.append("\n---\n_Past performance ≠ future results. Backtests are optimistic by nature "
             "(no liquidity, slippage during gaps, survivorship bias)._")

    path.write_text("\n".join(L), encoding="utf-8")


def write_chart(result: BacktestResult, benchmarks: dict[str, pd.Series], path: Path) -> None:
    """Equity curve PNG vs benchmarks. Requires matplotlib."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        log.warning("matplotlib not installed — skipping chart")
        return

    eq = pd.DataFrame([(p.date, p.total) for p in result.equity_curve],
                      columns=["date", "equity"]).set_index("date").sort_index()
    daily = eq["equity"].resample("D").last().ffill()

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 7), sharex=True,
                                    gridspec_kw={"height_ratios": [3, 1]})

    # Strategy
    ax1.plot(daily.index, daily.values, label="Strategy",
             linewidth=2.0, color="#1F4E78")

    # Benchmarks
    sip_active = bool(result.contributions) and result.config.sip_amount > 0
    for name, series in benchmarks.items():
        if series is None or series.empty:
            continue
        if sip_active:
            bench = benchmark_sip(series, result.contributions)
            label = f"{name} (SIP)"
        else:
            bench = benchmark_buy_and_hold(series, result.config.initial_capital)
            label = name
        if bench.empty:
            continue
        bench = bench.resample("D").last().ffill()
        # align
        bench = bench.reindex(daily.index, method="ffill")
        ax1.plot(bench.index, bench.values, label=label,
                 linewidth=1.0, alpha=0.7, linestyle="--")

    ax1.set_title(f"Equity Curve — {result.start} → {result.end}",
                  fontsize=13, fontweight="bold")
    ax1.set_ylabel("Equity (₹)")
    ax1.legend(loc="upper left")
    ax1.grid(True, alpha=0.3)
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"₹{x/1e5:.1f}L"))

    # Drawdown subplot
    cummax = daily.cummax()
    dd = (daily / cummax - 1) * 100
    ax2.fill_between(dd.index, dd.values, 0, color="red", alpha=0.3)
    ax2.plot(dd.index, dd.values, color="darkred", linewidth=0.8)
    ax2.set_ylabel("Drawdown (%)")
    ax2.set_xlabel("Date")
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(min(dd.min() - 2, -5), 1)

    plt.tight_layout()
    plt.savefig(path, dpi=120, bbox_inches="tight")
    plt.close(fig)
    log.info("Saved equity chart: %s", path)
