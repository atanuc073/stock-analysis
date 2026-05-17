"""Compare stop-loss stats between two backtests."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

PREV_XLSX = r"D:\MY_WORK\stock_analysis\reports\backtest\backtest_2018-01-01_2025-01-01_20260514_162210.xlsx"
NEW_XLSX = r"D:\MY_WORK\stock_analysis\reports\backtest\backtest_2018-01-01_2025-01-01_20260514_181724.xlsx"

def get_stats(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["All_Trades"]
    headers = [c.value for c in ws[1]]
    trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        trades.append(dict(zip(headers, row)))
    
    sells = [t for t in trades if t["Action"] == "SELL" and t["PnL_%"] is not None]
    stops = [t for t in sells if t["Exit_Type"] == "STOP_LOSS"]
    winners = [t for t in sells if t["PnL_%"] > 0]
    
    return {
        "total_trades": len(sells),
        "stops": len(stops),
        "stop_rate": len(stops)/len(sells)*100 if sells else 0,
        "stop_loss_abs": sum(t["PnL_Abs"] for t in stops if t["PnL_Abs"]),
        "win_rate": len(winners)/len(sells)*100 if sells else 0,
        "avg_win": sum(t["PnL_Abs"] for t in winners if t["PnL_Abs"]) / len(winners) if winners else 0,
        "avg_loss": sum(t["PnL_Abs"] for t in sells if t["PnL_Abs"] < 0) / len([t for t in sells if t["PnL_Abs"] < 0]) if [t for t in sells if t["PnL_Abs"] < 0] else 0
    }

prev = get_stats(PREV_XLSX)
new = get_stats(NEW_XLSX)

print(f"{'Metric':<20} {'Previous (16:22)':<20} {'New (18:17)':<20} {'Delta':<20}")
print("-" * 80)
for k in prev:
    p_val = prev[k]
    n_val = new[k]
    diff = n_val - p_val
    if isinstance(p_val, float):
        print(f"{k:<20} {p_val:>20.2f} {n_val:>20.2f} {diff:>+20.2f}")
    else:
        print(f"{k:<20} {p_val:>20,d} {n_val:>20,d} {diff:>+20,d}")
