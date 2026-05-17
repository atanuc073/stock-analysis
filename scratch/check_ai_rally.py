import openpyxl

PATH = r"D:\MY_WORK\stock_analysis\reports\backtest\backtest_2023-05-01_2024-12-31_20260514_193631.xlsx"

def check_ai_trades(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["All_Trades"]
    headers = [c.value for c in ws[1]]
    trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        trades.append(dict(zip(headers, row)))
    
    ai_symbols = ["NVDA", "AVGO", "MSFT", "SMCI", "MU", "VRT", "ARM", "GOOGL", "AMD"]
    found = [t for t in trades if t["Symbol"] in ai_symbols]
    
    print(f"AI Leader Trades Found: {len(found)}")
    for t in found:
        print(f"  {t['Symbol']} | {t['Action']} | {t['Date']} | PnL: {t['PnL_%']}% | Entry Score: {t['Score_Entry']}")

    # Top 5 winners overall
    sells = [t for t in trades if t["Action"] == "SELL" and t["PnL_%"] is not None]
    sells.sort(key=lambda x: x["PnL_%"], reverse=True)
    print("\nTop 5 Overall Winners:")
    for t in sells[:5]:
        print(f"  {t['Symbol']} | PnL: {t['PnL_%']}% | Days: {t['Days_Held']}")

check_ai_trades(PATH)
