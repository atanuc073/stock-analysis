"""Deep-dive analysis of stop-loss mechanics from backtest trades."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
import statistics

XLSX = r"D:\MY_WORK\stock_analysis\reports\backtest\backtest_2018-01-01_2025-01-01_20260514_162210.xlsx"
wb = openpyxl.load_workbook(XLSX)

ws = wb["All_Trades"]
headers = [c.value for c in ws[1]]
trades = []
for row in ws.iter_rows(min_row=2, values_only=True):
    trades.append(dict(zip(headers, row)))

# Pair BUY → SELL for each symbol occurrence
# Build timeline of events per symbol
sells = [t for t in trades if t["Action"] == "SELL" and t["PnL_%"] is not None]
stop_losses = [t for t in sells if t["Exit_Type"] == "STOP_LOSS"]

print("=" * 70)
print("STOP-LOSS ANATOMY")
print("=" * 70)
print(f"  Total stop-loss exits: {len(stop_losses)}")
print(f"  Total money lost: Rs{sum(t['PnL_Abs'] for t in stop_losses if t['PnL_Abs']):,.0f}")
print()

# ── 1. Distribution of loss severity ────────────────────────────────────
print("LOSS SEVERITY DISTRIBUTION:")
buckets = [
    ("Tiny (0 to -5%)",     lambda p: -5 < p <= 0),
    ("Small (-5 to -10%)",  lambda p: -10 < p <= -5),
    ("Medium (-10 to -15%)", lambda p: -15 < p <= -10),
    ("Large (-15 to -20%)", lambda p: -20 < p <= -15),
    ("Huge (< -20%)",       lambda p: p <= -20),
]
for label, fn in buckets:
    group = [t for t in stop_losses if fn(t["PnL_%"])]
    total = sum(t["PnL_Abs"] for t in group if t["PnL_Abs"])
    avg_days = statistics.mean([t["Days_Held"] for t in group if t["Days_Held"]]) if group else 0
    print(f"  {label:<25s}: {len(group):>4d} trades, Rs{total:>+10,.0f}, avg {avg_days:.0f} days")

# ── 2. Days held distribution ───────────────────────────────────────────
print()
print("DAYS HELD BEFORE STOP-LOSS:")
day_buckets = [
    ("1-7 days (first week)",    lambda d: 1 <= d <= 7),
    ("8-14 days (2nd week)",     lambda d: 8 <= d <= 14),
    ("15-30 days (month 1)",     lambda d: 15 <= d <= 30),
    ("31-60 days (month 2)",     lambda d: 31 <= d <= 60),
    ("61-90 days (quarter 1)",   lambda d: 61 <= d <= 90),
    ("91-180 days (6 months)",   lambda d: 91 <= d <= 180),
    (">180 days",                lambda d: d > 180),
]
for label, fn in day_buckets:
    group = [t for t in stop_losses if t["Days_Held"] and fn(t["Days_Held"])]
    if not group:
        print(f"  {label:<30s}: 0 trades")
        continue
    pnls = [t["PnL_%"] for t in group]
    total = sum(t["PnL_Abs"] for t in group if t["PnL_Abs"])
    print(f"  {label:<30s}: {len(group):>4d} trades, avg {statistics.mean(pnls):>+.1f}%, Rs{total:>+10,.0f}")

# ── 3. Stop distance at entry (how tight?) ───────────────────────────────
# We can infer the initial stop distance from the final PnL% of stop-loss trades
# because stop-loss fires at stop_price, so PnL% ≈ stop distance
print()
print("INITIAL STOP TIGHTNESS (inferred from stop-loss PnL%):")
print("  (PnL% at stop-loss ≈ how far the stop was from entry)")
sl_pnls = sorted([t["PnL_%"] for t in stop_losses])
print(f"  Tightest stop hit: {max(sl_pnls):+.1f}%")
print(f"  Widest stop hit:   {min(sl_pnls):+.1f}%")
print(f"  Mean:              {statistics.mean(sl_pnls):+.1f}%")
print(f"  Median:            {statistics.median(sl_pnls):+.1f}%")
print(f"  Std Dev:           {statistics.stdev(sl_pnls):.1f}%")

# How many were stopped at a gain (regime tightened stop)?
wins_on_stop = [t for t in stop_losses if t["PnL_%"] > 0]
print(f"\n  Stopped out at a GAIN: {len(wins_on_stop)} trades ({len(wins_on_stop)/len(stop_losses)*100:.1f}%)")
if wins_on_stop:
    for t in wins_on_stop[:8]:
        print(f"    {t['Symbol']:<15s} {t['PnL_%']:>+.1f}% after {t['Days_Held']} days")

# ── 4. Quick stops that would have recovered ─────────────────────────────
print()
print("=" * 70)
print("QUICK STOP-LOSSES (≤30 days) — DETAIL")
print("=" * 70)
quick = [t for t in stop_losses if t["Days_Held"] and t["Days_Held"] <= 30]
quick.sort(key=lambda t: t["PnL_Abs"] or 0)
print(f"  {'Symbol':<18s} {'Mkt':>3s} {'PnL%':>8s} {'PnL Abs':>12s} {'Days':>5s} {'Score':>6s} {'Sector':<20s}")
print(f"  {'-'*18} {'-'*3} {'-'*8} {'-'*12} {'-'*5} {'-'*6} {'-'*20}")
for t in quick[:25]:
    print(f"  {t['Symbol']:<18s} {t['Mkt']:>3s} {t['PnL_%']:>+8.1f} Rs{t['PnL_Abs']:>+10,.0f} "
          f"{t['Days_Held'] or 0:>5} {t['Score_Entry'] or 0:>6.1f} {(t['Sector'] or '')[:20]:<20s}")

print(f"\n  Total quick-stop losses: Rs{sum(t['PnL_Abs'] for t in quick if t['PnL_Abs']):,.0f}")

# ── 5. Repeat offenders (same stock stopped multiple times) ──────────────
print()
print("=" * 70)
print("REPEAT OFFENDERS (stopped out 2+ times on same stock)")
print("=" * 70)
from collections import Counter
sym_counts = Counter(t["Symbol"] for t in stop_losses)
repeats = {s: c for s, c in sym_counts.items() if c >= 2}
if repeats:
    print(f"  {'Symbol':<18s} {'Stop Count':>10s} {'Total Lost':>12s}")
    print(f"  {'-'*18} {'-'*10} {'-'*12}")
    for sym in sorted(repeats, key=lambda s: sum(t["PnL_Abs"] for t in stop_losses if t["Symbol"] == s and t["PnL_Abs"])):
        total = sum(t["PnL_Abs"] for t in stop_losses if t["Symbol"] == sym and t["PnL_Abs"])
        print(f"  {sym:<18s} {repeats[sym]:>10d} Rs{total:>+10,.0f}")

# ── 6. Regime at entry for stop-loss trades ──────────────────────────────
print()
print("=" * 70)
print("WHAT REGIME WERE YOU IN WHEN YOU BOUGHT THESE LOSERS?")
print("=" * 70)
# Match BUY trades to their SELL stop-loss trades by symbol + timing
buys = [t for t in trades if t["Action"] == "BUY"]

# Build BUY context lookup: for each symbol, list of BUY entries
buy_lookup = {}
for b in buys:
    if b["Symbol"] not in buy_lookup:
        buy_lookup[b["Symbol"]] = []
    buy_lookup[b["Symbol"]].append(b)

# For stop-loss sells, find the matching BUY (most recent before the sell date)
regime_stats = {}
for sl in stop_losses:
    sym = sl["Symbol"]
    possible_buys = buy_lookup.get(sym, [])
    # Find BUY that happened before this SELL, closest in time
    matching = [b for b in possible_buys if b["Date"] <= sl["Date"]]
    if not matching:
        continue
    buy = matching[-1]  # most recent BUY before SELL
    regime = buy.get("Reason", "")
    # Extract regime from the reason field if available
    # Try to get regime from entry context fields
    # Actually the regime_label_at_entry is in BTTrade but not in Excel
    # Let's just use the entry score
    score = buy.get("Score_Entry", 0)
    score_bucket = f"{int(score // 5) * 5}-{int(score // 5) * 5 + 5}" if score else "unknown"
    if score_bucket not in regime_stats:
        regime_stats[score_bucket] = {"count": 0, "total_loss": 0}
    regime_stats[score_bucket]["count"] += 1
    regime_stats[score_bucket]["total_loss"] += sl["PnL_Abs"] or 0

print("  Entry Score Bucket → Stop-Loss Outcomes:")
print(f"  {'Score Bucket':<15s} {'Stops':>6s} {'Total Lost':>12s} {'Avg Lost':>10s}")
print(f"  {'-'*15} {'-'*6} {'-'*12} {'-'*10}")
for bucket in sorted(regime_stats.keys()):
    d = regime_stats[bucket]
    avg = d["total_loss"] / d["count"] if d["count"] else 0
    print(f"  {bucket:<15s} {d['count']:>6d} Rs{d['total_loss']:>+10,.0f} Rs{avg:>+8,.0f}")

# ── 7. Current stop-loss config summary ──────────────────────────────────
print()
print("=" * 70)
print("YOUR CURRENT STOP-LOSS CONFIGURATION")
print("=" * 70)
print(f"  ATR multiplier:        3.0x  (stop = entry - 3×ATR)")
print(f"  Hard stop floor:       -18%  (never risk more than 18%)")
print(f"  Confirm bars:          2     (must close below stop 2 consecutive days)")
print(f"  Hard stop buffer:      2.5%  (if price drops 2.5% BELOW stop, fire immediately)")
print(f"  Trailing stop:         18%   (off peak, after all tiers triggered)")
print(f"  Time stop:             365d  (exit if return between -5% and +10%)")
print()
print("  REGIME STOP TIGHTENING:")
print(f"    BULL:          1.0x  (no change)")
print(f"    NEUTRAL_BULL:  1.0x  (no change)")
print(f"    NEUTRAL:       0.85x (stop 15% tighter)")
print(f"    CAUTIOUS:      0.65x (stop 35% tighter)")
print(f"    BEAR:          0.50x (stop 50% tighter)")
