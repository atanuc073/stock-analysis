"""Analyze backtest trades: are we buying at tops and losing money?"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
import statistics

XLSX = r"D:\MY_WORK\stock_analysis\reports\backtest\backtest_2018-01-01_2025-01-01_20260514_162210.xlsx"

wb = openpyxl.load_workbook(XLSX)

# ── 1. Parse All Trades ──────────────────────────────────────────────────
ws = wb["All_Trades"]
headers = [c.value for c in ws[1]]
trades = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    trades.append(d)

# Pair BUYs with SELLs
buys = {t["Symbol"]: [] for t in trades}
for t in trades:
    buys[t["Symbol"]].append(t)

# Collect closed trades (SELLs with PnL)
sells = [t for t in trades if t["Action"] == "SELL" and t["PnL_%"] is not None]
print(f"Total closed trades: {len(sells)}")
print(f"  Winners: {sum(1 for s in sells if s['PnL_%'] > 0)}")
print(f"  Losers:  {sum(1 for s in sells if s['PnL_%'] <= 0)}")
print()

# ── 2. Parse TopChase data ───────────────────────────────────────────────
print("=" * 70)
print("TOP-CHASE SUMMARY (from report)")
print("=" * 70)
ws_tc = wb["TopChase_Summary"]
for row in ws_tc.iter_rows(min_row=1, values_only=True):
    print(f"  {row[0]:40s} : {row[1]}")

print()
print("=" * 70)
print("FORWARD RETURNS BY DISTANCE FROM 52-WEEK HIGH")
print("=" * 70)
ws_hi = wb["TopChase_By_52wHigh"]
hi_headers = [c.value for c in ws_hi[1]]
print(f"  {'Bucket':<25s} {'Trades':>7s} {'Avg30':>8s} {'Med30':>8s} {'Win30':>8s} {'Avg90':>8s} {'Win90':>8s}")
print(f"  {'-'*25} {'-'*7} {'-'*8} {'-'*8} {'-'*8} {'-'*8} {'-'*8}")
for row in ws_hi.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    print(f"  {str(row[0]):<25s} {row[1]:>7} {row[2]:>8.2f} {row[3]:>8.2f} {row[4]:>8.1f} {row[5]:>8.2f} {row[6]:>8.1f}")

print()
print("=" * 70)
print("FORWARD RETURNS BY EXTENSION ABOVE 200DMA")
print("=" * 70)
ws_ext = wb["TopChase_By_Extension"]
ext_headers = [c.value for c in ws_ext[1]]
print(f"  {'Bucket':<15s} {'Trades':>7s} {'Avg30':>8s} {'Med30':>8s} {'Win30':>8s} {'Avg90':>8s} {'Win90':>8s}")
print(f"  {'-'*15} {'-'*7} {'-'*8} {'-'*8} {'-'*8} {'-'*8} {'-'*8}")
for row in ws_ext.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    print(f"  {str(row[0]):<15s} {row[1]:>7} {row[2]:>8.2f} {row[3]:>8.2f} {row[4]:>8.1f} {row[5]:>8.2f} {row[6]:>8.1f}")

# ── 3. Analyze SELL trades by exit type ──────────────────────────────────
print()
print("=" * 70)
print("P&L BREAKDOWN BY EXIT TYPE")
print("=" * 70)

by_exit = {}
for s in sells:
    et = s["Exit_Type"] or "UNKNOWN"
    if et not in by_exit:
        by_exit[et] = []
    by_exit[et].append(s)

print(f"  {'Exit Type':<18s} {'Count':>6s} {'Avg PnL%':>9s} {'Med PnL%':>9s} {'WinRate':>8s} {'Total PnL':>12s} {'Avg Days':>9s}")
print(f"  {'-'*18} {'-'*6} {'-'*9} {'-'*9} {'-'*8} {'-'*12} {'-'*9}")
for et in sorted(by_exit, key=lambda x: -len(by_exit[x])):
    group = by_exit[et]
    pnls = [g["PnL_%"] for g in group]
    abs_pnls = [g["PnL_Abs"] for g in group if g["PnL_Abs"] is not None]
    days = [g["Days_Held"] for g in group if g["Days_Held"] is not None and g["Days_Held"] > 0]
    wins = sum(1 for p in pnls if p > 0)
    print(f"  {et:<18s} {len(group):>6d} {statistics.mean(pnls):>+9.2f} {statistics.median(pnls):>+9.2f} "
          f"{wins/len(group)*100:>7.1f}% {sum(abs_pnls):>12,.0f} {statistics.mean(days) if days else 0:>9.0f}")

# ── 4. Stop loss analysis: are these from top-chasing? ───────────────────
print()
print("=" * 70)
print("STOP-LOSS DEEP DIVE (are stop losses from buying at tops?)")
print("=" * 70)

stop_losses = [s for s in sells if s["Exit_Type"] == "STOP_LOSS"]
other_losses = [s for s in sells if s["Exit_Type"] != "STOP_LOSS" and s["PnL_%"] <= 0]

print(f"\n  Stop-Loss trades: {len(stop_losses)}")
if stop_losses:
    sl_pnls = [s["PnL_%"] for s in stop_losses]
    sl_abs = [s["PnL_Abs"] for s in stop_losses if s["PnL_Abs"] is not None]
    sl_days = [s["Days_Held"] for s in stop_losses if s["Days_Held"] and s["Days_Held"] > 0]
    print(f"  Avg PnL%:     {statistics.mean(sl_pnls):+.2f}%")
    print(f"  Median PnL%:  {statistics.median(sl_pnls):+.2f}%")
    print(f"  Total $ Lost: ₹{sum(sl_abs):,.0f}")
    print(f"  Avg Days Held: {statistics.mean(sl_days):.0f}")

    # Quick vs slow stops
    quick = [s for s in stop_losses if s["Days_Held"] and s["Days_Held"] <= 30]
    slow = [s for s in stop_losses if s["Days_Held"] and s["Days_Held"] > 30]
    print(f"\n  Quick stops (≤30 days): {len(quick)}")
    if quick:
        print(f"    Avg PnL%: {statistics.mean([s['PnL_%'] for s in quick]):+.2f}%")
        print(f"    Total Lost: ₹{sum(s['PnL_Abs'] for s in quick if s['PnL_Abs']):,.0f}")
    print(f"  Slow stops (>30 days): {len(slow)}")
    if slow:
        print(f"    Avg PnL%: {statistics.mean([s['PnL_%'] for s in slow]):+.2f}%")
        print(f"    Total Lost: ₹{sum(s['PnL_Abs'] for s in slow if s['PnL_Abs']):,.0f}")

# ── 5. Score at entry for winners vs losers ──────────────────────────────
print()
print("=" * 70)
print("ENTRY SCORE: WINNERS vs LOSERS")
print("=" * 70)

# Match BUY scores to SELL outcomes
winners = [s for s in sells if s["PnL_%"] > 0]
losers = [s for s in sells if s["PnL_%"] <= 0]

win_scores = [s["Score_Entry"] for s in winners if s["Score_Entry"] is not None]
lose_scores = [s["Score_Entry"] for s in losers if s["Score_Entry"] is not None]

if win_scores and lose_scores:
    print(f"  Winners avg entry score: {statistics.mean(win_scores):.1f} (n={len(win_scores)})")
    print(f"  Losers avg entry score:  {statistics.mean(lose_scores):.1f} (n={len(lose_scores)})")
    print(f"  Difference: {statistics.mean(win_scores) - statistics.mean(lose_scores):+.1f}")

# ── 6. Biggest losers detail ─────────────────────────────────────────────
print()
print("=" * 70)
print("TOP 15 BIGGEST LOSSES (absolute)")
print("=" * 70)
biggest_losses = sorted(sells, key=lambda s: s["PnL_Abs"] or 0)[:15]
print(f"  {'Symbol':<20s} {'Mkt':>3s} {'PnL%':>8s} {'PnL Abs':>12s} {'Days':>5s} {'Exit':>15s} {'Score':>6s} {'Sector':<20s}")
print(f"  {'-'*20} {'-'*3} {'-'*8} {'-'*12} {'-'*5} {'-'*15} {'-'*6} {'-'*20}")
for s in biggest_losses:
    print(f"  {s['Symbol']:<20s} {s['Mkt']:>3s} {s['PnL_%']:>+8.1f} ₹{s['PnL_Abs']:>+10,.0f} "
          f"{s['Days_Held'] or 0:>5} {s['Exit_Type'] or '':>15s} {s['Score_Entry'] or 0:>6.1f} {(s['Sector'] or '')[:20]:<20s}")

total_loss_from_top15 = sum(s["PnL_Abs"] for s in biggest_losses if s["PnL_Abs"])
total_all_losses = sum(s["PnL_Abs"] for s in sells if s["PnL_Abs"] and s["PnL_Abs"] < 0)
total_all_wins = sum(s["PnL_Abs"] for s in sells if s["PnL_Abs"] and s["PnL_Abs"] > 0)

print(f"\n  Top 15 losses total: ₹{total_loss_from_top15:,.0f}")
print(f"  ALL losses total:   ₹{total_all_losses:,.0f}")
print(f"  ALL wins total:     ₹{total_all_wins:+,.0f}")
print(f"  Net P&L:            ₹{total_all_wins + total_all_losses:+,.0f}")
print(f"  Top 15 losses as % of all losses: {total_loss_from_top15/total_all_losses*100:.1f}%")

# ── 7. Sector breakdown for losses ──────────────────────────────────────
print()
print("=" * 70)
print("LOSSES BY SECTOR")
print("=" * 70)
sector_losses = {}
for s in sells:
    if s["PnL_Abs"] and s["PnL_Abs"] < 0:
        sec = s["Sector"] or "Unknown"
        if sec not in sector_losses:
            sector_losses[sec] = {"count": 0, "total": 0, "pnls": []}
        sector_losses[sec]["count"] += 1
        sector_losses[sec]["total"] += s["PnL_Abs"]
        sector_losses[sec]["pnls"].append(s["PnL_%"])

print(f"  {'Sector':<25s} {'Losses':>7s} {'Total Lost':>12s} {'Avg PnL%':>9s}")
print(f"  {'-'*25} {'-'*7} {'-'*12} {'-'*9}")
for sec in sorted(sector_losses, key=lambda x: sector_losses[x]["total"]):
    d = sector_losses[sec]
    print(f"  {sec[:25]:<25s} {d['count']:>7d} ₹{d['total']:>+11,.0f} {statistics.mean(d['pnls']):>+9.2f}%")
